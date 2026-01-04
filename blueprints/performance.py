#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
绩效管理模块
负责绩效数据上传、统计、导出等功能
"""
import os
import re
import sqlite3
from datetime import datetime
from typing import Dict, List, Tuple

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from openpyxl import Workbook
from werkzeug.utils import secure_filename

from config.settings import APP_TITLE, UPLOAD_DIR, EXPORT_DIR
from models.database import get_db
from .decorators import login_required, role_required
from .helpers import require_user_id, build_department_filter, log_import_operation

# 创建 Blueprint
performance_bp = Blueprint('performance', __name__, url_prefix='/performance')


# ==================== 常量定义 ====================

ALLOWED_EXTS = {"pdf"}

# PDF解析正则表达式
HEADER_PERIOD_RE = re.compile(r"考核周期\s+(\d{4})[/-](\d{1,2})[/-](\d{1,2})")
ROW_RE = re.compile(
    r"^\s*\d+\s+"
    r"(?P<emp_no>\d{1,20})\s+"
    r"(?P<name>[\u4e00-\u9fa5A-Za-z·.\-]{1,30})\s+"
    r"(?P<grade>A|B\+|B|C|D)\s+"
    r"(?P<score>\d+(?:\.\d+)?)\s+"
)

# 默认档位映射
DEFAULT_GRADE_MAP = {"A": 5, "B+": 4, "B": 3, "C": 2, "D": 1, "不进行": 0}

# 默认季度等级
DEFAULT_QUARTER_GRADES = ["优秀", "良好", "称职", "待改进", "不合格"]
DEFAULT_QUARTER_DEFAULT = "称职"
DEFAULT_QUARTER_COLORS = {
    "优秀": "#E6F4EA",
    "良好": "#E3F2FD",
    "称职": "#FFF8E1",
    "待改进": "#FCE4EC",
    "不合格": "#FFEBEE",
}


# ==================== 辅助函数 ====================

def extract_text_from_pdf(pdf_path):
    """从PDF提取文本"""
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text(x_tolerance=1, y_tolerance=1) or ""
                text += "\n"
        if text.strip():
            return text
    except Exception as exc:
        print("pdfplumber failed:", exc)

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
        if text.strip():
            return text
    except Exception as exc:
        print("PyPDF2 failed:", exc)
    raise RuntimeError("无法从PDF提取文本，请确认PDF包含文本层并非扫描件。")


def parse_pdf_text(text: str):
    """解析PDF文本，提取绩效数据"""
    year, month = None, None
    m = HEADER_PERIOD_RE.search(text)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
    rows = []
    for line in text.splitlines():
        line = line.strip()
        m = ROW_RE.match(line)
        if not m:
            continue
        d = m.groupdict()
        rows.append({
            "emp_no": d["emp_no"],
            "name": d["name"],
            "grade": d["grade"],
            "score": float(d["score"]),
        })
    return year, month, rows


def emp_sort_key(emp_no: str) -> int:
    """员工编号排序键"""
    digits = re.sub(r"\D", "", emp_no or "")
    return int(digits) if digits else 10**9


def ym_to_int(y, m):
    """年月转整数（用于区间比较）"""
    return int(f"{int(y):04d}{int(m):02d}")


def build_yearly_matrix(uid: int, year: int, accessible_user_ids=None):
    """构建年度绩效矩阵

    注意: accessible_user_ids参数已废弃，保留只为向后兼容
    """
    months = [(year, m) for m in range(1, 13)]
    conn = get_db()
    cur = conn.cursor()

    # 使用新的部门过滤机制
    where_clause, join_clause, dept_params = build_department_filter('pr')

    cur.execute(
        f"""
        SELECT pr.emp_no, pr.name, pr.year, pr.month, pr.score, pr.grade
        FROM performance_records pr
        {join_clause}
        WHERE {where_clause} AND pr.year=?
        ORDER BY CAST(pr.emp_no as INTEGER), pr.month
        """,
        dept_params + [year],
    )
    rows = cur.fetchall()

    grade_choices = sorted({row["grade"] for row in rows if row["grade"]})
    by_emp: Dict[str, Dict[str, object]] = {}
    for r in rows:
        key = r["emp_no"]
        bucket = by_emp.setdefault(key, {"emp_no": r["emp_no"], "name": r["name"], "items": {}})
        bucket["items"][(r["year"], r["month"])] = {"score": r["score"], "grade": r["grade"]}

    data = []
    for bucket in by_emp.values():
        detail = []
        scores = []
        for (y, m) in months:
            item = bucket["items"].get((y, m))
            if item:
                detail.append(item)
                if item.get("score") is not None:
                    scores.append(item["score"])
            else:
                detail.append(None)
        avg_score = round(sum(scores) / len(scores), 2) if scores else None
        data.append({"emp_no": bucket["emp_no"], "name": bucket["name"], "detail": detail, "avg_score": avg_score})

    return data, months, grade_choices


def filter_sort_yearly_data(data, grade_filter_set, min_score_filter, sort):
    """过滤和排序年度数据"""
    filtered = []
    for row in data:
        grades = [cell["grade"] for cell in row["detail"] if cell and cell.get("grade")]
        grade_match = not grade_filter_set or any(g in grade_filter_set for g in grades)
        score_match = min_score_filter is None or (row["avg_score"] is not None and row["avg_score"] >= min_score_filter)
        if grade_match and score_match:
            filtered.append(row)

    if sort == "avg_desc":
        filtered.sort(key=lambda r: (r["avg_score"] is None, -(r["avg_score"] or 0), emp_sort_key(r["emp_no"]), r["name"]))
    elif sort == "avg_asc":
        filtered.sort(key=lambda r: (r["avg_score"] is None, (r["avg_score"] or 10**9), emp_sort_key(r["emp_no"]), r["name"]))
    else:
        filtered.sort(key=lambda r: (emp_sort_key(r["emp_no"]), r["name"]))
    return filtered


def build_calculator_dataset(uid: int, year: int, mapping: Dict[str, float]):
    """构建绩效计算器数据集"""
    months = [(year, m) for m in range(1, 13)]
    conn = get_db()
    cur = conn.cursor()

    # 使用部门过滤机制
    where_clause, join_clause, dept_params = build_department_filter('pr')

    try:
        cur.execute(
            f"""
            SELECT pr.emp_no, pr.name, pr.year, pr.month, pr.grade
            FROM performance_records pr
            {join_clause}
            WHERE {where_clause} AND pr.year=?
            ORDER BY CAST(pr.emp_no as INTEGER), pr.month
            """,
            dept_params + [year],
        )
    except sqlite3.OperationalError:
        cur.execute(
            f"""
            SELECT pr.emp_no, pr.name, pr.year, pr.month, pr.grade
            FROM performance_records pr
            {join_clause}
            WHERE {where_clause} AND pr.year=?
            ORDER BY pr.month
            """,
            dept_params + [year],
        )
    rows = cur.fetchall()

    by_emp: Dict[str, Dict[str, object]] = {}
    for r in rows:
        bucket = by_emp.setdefault(r["emp_no"], {"emp_no": r["emp_no"], "name": r["name"], "items": {}})
        bucket["items"][(r["year"], r["month"])] = {"grade": r["grade"]}

    data = []
    for bucket in by_emp.values():
        detail = []
        total = 0.0
        for (y, m) in months:
            item = bucket["items"].get((y, m))
            if item and item.get("grade"):
                grade = item["grade"]
                value = mapping.get(grade, 0.0)
                detail.append({"grade": grade, "value": value})
                total += value
            else:
                detail.append(None)
        data.append({"emp_no": bucket["emp_no"], "name": bucket["name"], "detail": detail, "total": round(total, 2)})

    return data, months


def sort_calculator_data(data, sort):
    """排序计算器数据"""
    if sort == "total_desc":
        data.sort(key=lambda r: (-r["total"], emp_sort_key(r["emp_no"]), r["name"]))
    elif sort == "total_asc":
        data.sort(key=lambda r: (r["total"], emp_sort_key(r["emp_no"]), r["name"]))
    else:
        data.sort(key=lambda r: (emp_sort_key(r["emp_no"]), r["name"]))
    return data


def get_or_init_grade_map(uid: int = None):
    """获取或初始化档位映射（全局配置，不再按用户隔离）"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT grade, value FROM grade_map")
    rows = cur.fetchall()
    if not rows:
        # 初始化默认档位映射
        for grade, value in DEFAULT_GRADE_MAP.items():
            cur.execute("INSERT OR REPLACE INTO grade_map(grade, value) VALUES(?,?)", (grade, float(value)))
        conn.commit()
        cur.execute("SELECT grade, value FROM grade_map")
        rows = cur.fetchall()
    return {row["grade"]: float(row["value"]) for row in rows}


def normalize_color(value: str, fallback: str) -> str:
    """标准化颜色值"""
    value = (value or "").strip()
    if not value:
        return fallback
    if re.fullmatch(r"#[0-9a-fA-F]{6}", value):
        return value.upper()
    if re.fullmatch(r"#[0-9a-fA-F]{3}", value):
        return "#" + "".join(ch * 2 for ch in value[1:]).upper()
    return fallback


def set_quarter_grade_options(uid: int, entries: List[Tuple[str, str]], default_grade: str) -> None:
    """设置季度等级选项（全局配置，不再按用户隔离）"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM quarter_grade_options")
    for idx, (grade, color) in enumerate(entries):
        cur.execute(
            """
            INSERT INTO quarter_grade_options(grade, display_order, is_default, color)
            VALUES (?, ?, ?, ?)
            """,
            (grade, idx, 1 if grade == default_grade else 0, color),
        )
    conn.commit()


def get_quarter_grade_options(uid: int):
    """获取季度等级选项（全局配置，不再按用户隔离）"""
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT grade, display_order, is_default, color FROM quarter_grade_options ORDER BY display_order"
    )
    rows = cur.fetchall()
    if not rows:
        defaults = [(grade, DEFAULT_QUARTER_COLORS.get(grade, "#FFFFFF")) for grade in DEFAULT_QUARTER_GRADES]
        set_quarter_grade_options(uid, defaults, DEFAULT_QUARTER_DEFAULT)
        cur.execute(
            "SELECT grade, display_order, is_default, color FROM quarter_grade_options ORDER BY display_order"
        )
        rows = cur.fetchall()
    return [
        {
            "grade": row["grade"],
            "is_default": bool(row["is_default"]),
            "color": row["color"] or DEFAULT_QUARTER_COLORS.get(row["grade"], "#FFFFFF"),
        }
        for row in rows
    ]


def list_employees():
    """列出员工（用于季度数据集）"""
    from .personnel import list_personnel
    return list_personnel()


def build_quarter_dataset(uid: int, year: int, options=None):
    """构建季度绩效数据集"""
    options = options or get_quarter_grade_options(uid)
    default_grade = next((opt["grade"] for opt in options if opt["is_default"]), options[0]["grade"])

    try:
        emps = list_employees()
    except Exception:
        emps = []

    conn = get_db()
    cur = conn.cursor()

    # 使用部门过滤机制
    where_clause, join_clause, dept_params = build_department_filter('pr')
    cur.execute(
        f"""
        SELECT pr.emp_no, pr.name, pr.month, pr.score, pr.grade
        FROM performance_records pr
        {join_clause}
        WHERE {where_clause} AND pr.year=?
        ORDER BY pr.month
        """,
        dept_params + [year]
    )
    perf_rows = cur.fetchall()

    # quarter_overrides 不再使用 user_id 过滤
    cur.execute("SELECT emp_no, quarter, grade FROM quarter_overrides WHERE year=?", (year,))
    overrides = {(row["emp_no"], row["quarter"]): row["grade"] for row in cur.fetchall()}

    month_groups = {1: (1, 2, 3), 2: (4, 5, 6), 3: (7, 8, 9), 4: (10, 11, 12)}
    perf = {}
    for row in perf_rows:
        perf.setdefault(row["emp_no"], {})[row["month"]] = {"score": row["score"], "grade": row["grade"], "name": row["name"]}

    if not emps:
        seen = set()
        derived = []
        for emp_no, months_map in perf.items():
            if emp_no in seen:
                continue
            name = ""
            for row in perf_rows:
                if row["emp_no"] == emp_no:
                    name = row["name"]
                    break
            derived.append({"emp_no": emp_no, "name": name})
            seen.add(emp_no)
        emps = derived

    grade_color_lookup = {opt["grade"]: opt["color"] for opt in options}
    data = []
    for emp in emps:
        emp_no = emp["emp_no"] if isinstance(emp, dict) else emp["emp_no"]
        name = emp["name"] if isinstance(emp, dict) else emp["name"]
        row = {"emp_no": emp_no, "name": name, "q": {}}
        for quarter, months_tuple in month_groups.items():
            values = [perf.get(emp_no, {}).get(m) for m in months_tuple]
            scores = [v["score"] for v in values if v and v.get("score") is not None]
            score_avg = round(sum(scores) / len(scores), 2) if len(scores) == len(values) and values else ""
            grade_value = overrides.get((emp_no, quarter)) or default_grade
            row["q"][quarter] = {"score": score_avg, "grade": grade_value, "color": grade_color_lookup.get(grade_value, "")}
        data.append(row)

    return options, default_grade, data


def filter_quarter_data(data, grade_filter_set, min_score_filter):
    """过滤季度数据"""
    filtered = []
    for row in data:
        row_grades = [row["q"][q].get("grade") for q in (1, 2, 3, 4) if row["q"][q].get("grade")]
        row_scores = [row["q"][q].get("score") for q in (1, 2, 3, 4) if isinstance(row["q"][q].get("score"), (int, float))]
        grade_match = not grade_filter_set or any(g in grade_filter_set for g in row_grades)
        score_match = min_score_filter is None or any(s >= min_score_filter for s in row_scores)
        if grade_match and score_match:
            filtered.append(row)
    filtered.sort(key=lambda r: (emp_sort_key(r["emp_no"]), r["name"]))
    return filtered


# ==================== 路由定义 ====================

@performance_bp.route('/')
@login_required
def index():
    """绩效管理主控制台"""
    feature_cards = [
        {
            "title": "年度总览",
            "description": "查看年度绩效情况并导出数据。",
            "endpoint": "performance.records",
        },
        {
            "title": "上传PDF",
            "description": "导入月度绩效PDF，自动解析成绩并保存。",
            "endpoint": "performance.upload",
        },
        {
            "title": "区间统计",
            "description": "选择自定义时间区间，对比不同员工的绩效表现。",
            "endpoint": "performance.range_view",
        },
        {
            "title": "绩效计算器",
            "description": "使用自定义档位快速计算绩效分值。",
            "endpoint": "performance.calculator",
        },
        {
            "title": "季度绩效",
            "description": "配置季度档位，并对个人季度结果做人工修正。",
            "endpoint": "performance.quarters",
        },
    ]
    return render_template(
        "performance.html",
        title=f"绩效中心 | {APP_TITLE}",
        feature_cards=feature_cards,
    )


@performance_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    """上传并解析月度绩效PDF"""
    now = datetime.now()
    year_options = list(range(now.year - 5, now.year + 2))
    selected_year = request.args.get("year", type=int) or now.year
    selected_month = request.args.get("month", type=int) or now.month

    if not (1 <= selected_month <= 12):
        selected_month = now.month

    if not (2000 <= selected_year <= 2100):
        selected_year = now.year

    if selected_year not in year_options:
        year_options.append(selected_year)
        year_options.sort()

    if request.method == "POST":
        year_raw = (request.form.get("target_year") or "").strip()
        month_raw = (request.form.get("target_month") or "").strip()
        try:
            selected_year = int(year_raw)
            selected_month = int(month_raw)
        except ValueError:
            flash("请选择有效的年份和月份。", "warning")
            return redirect(url_for("performance.upload"))

        if selected_year not in year_options:
            year_options.append(selected_year)
            year_options.sort()

        if not (2000 <= selected_year <= 2100) or not (1 <= selected_month <= 12):
            flash("请选择有效的年份和月份。", "warning")
            return redirect(url_for("performance.upload"))

        force_import = request.form.get("force_import") == "1"
        pending_filename = (request.form.get("pending_filename") or "").strip()
        pending_filename = os.path.basename(pending_filename) if pending_filename else ""
        filename = ""

        if pending_filename:
            filename = pending_filename
            save_path = os.path.join(UPLOAD_DIR, filename)
            if not os.path.isfile(save_path):
                flash("原文件已不存在，请重新上传。", "danger")
                return redirect(url_for("performance.upload"))
        else:
            file_obj = request.files.get("file")
            if not file_obj or file_obj.filename == "":
                flash("请选择PDF文件", "warning")
                return redirect(url_for("performance.upload"))
            ext = file_obj.filename.rsplit(".", 1)[-1].lower()
            if ext not in ALLOWED_EXTS:
                flash("仅支持PDF文件", "warning")
                return redirect(url_for("performance.upload"))
            filename = datetime.now().strftime("%Y%m%d_%H%M%S_") + secure_filename(file_obj.filename)
            save_path = os.path.join(UPLOAD_DIR, filename)
            file_obj.save(save_path)

        try:
            text = extract_text_from_pdf(save_path)
            parsed_year, parsed_month, rows = parse_pdf_text(text)
            if not rows:
                flash("未识别到任何有效行，请确认PDF格式。", "danger")
                if not pending_filename and os.path.exists(save_path):
                    os.remove(save_path)
                return redirect(url_for("performance.upload"))

            mismatch = (
                parsed_year is not None
                and parsed_month is not None
                and (parsed_year != selected_year or parsed_month != selected_month)
            )

            if mismatch and not force_import:
                flash(
                    f"识别到的考核周期为 {parsed_year} 年 {parsed_month} 月，与选择的 {selected_year} 年 {selected_month} 月不一致。如确认以选择的周期导入，请再次提交确认。",
                    "warning",
                )
                return render_template(
                    "upload.html",
                    title=f"上传PDF | {APP_TITLE}",
                    year_options=year_options,
                    selected_year=selected_year,
                    selected_month=selected_month,
                    confirm_ctx={
                        "parsed_year": parsed_year,
                        "parsed_month": parsed_month,
                        "filename": filename,
                    },
                )

            uid = require_user_id()
            conn = get_db()
            cur = conn.cursor()

            # 使用部门过滤机制获取员工花名册
            where_clause, join_clause, dept_params = build_department_filter()
            cur.execute(
                f"""
                SELECT emp_no
                FROM employees
                {join_clause}
                WHERE {where_clause}
                """,
                dept_params
            )
            roster = {record[0] for record in cur.fetchall()}
            if not roster:
                flash("人员档案为空，请先新增人员。", "warning")
                return redirect(url_for("personnel.index"))

            filtered = [r for r in rows if r["emp_no"] in roster]
            skipped = len(rows) - len(filtered)

            if not filtered:
                flash("识别的记录均不在花名册中，未导入任何数据。", "warning")
                return redirect(url_for("performance.upload"))

            imported = 0
            for row in filtered:
                cur.execute(
                    """
                    INSERT INTO performance_records(emp_no, name, year, month, score, grade, src_file, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(emp_no, year, month) DO UPDATE SET
                        name=excluded.name,
                        score=excluded.score,
                        grade=excluded.grade,
                        src_file=excluded.src_file
                    """,
                    (
                        row["emp_no"],
                        row["name"],
                        selected_year,
                        selected_month,
                        row["score"],
                        row["grade"],
                        filename,
                        uid,
                    ),
                )
                imported += 1
            conn.commit()

            # 记录导入操作日志
            log_import_operation(
                module='performance',
                operation='import',
                file_name=filename,
                total_rows=len(rows),
                success_rows=imported,
                failed_rows=0,
                skipped_rows=skipped,
                import_details={
                    'year': selected_year,
                    'month': selected_month,
                    'imported': imported,
                    'skipped_not_in_roster': skipped,
                    'source_file': filename
                }
            )

            msg = f"导入成功：{selected_year}年{selected_month}月 共 {imported} 条。"
            if skipped:
                msg += f"（已跳过 {skipped} 条不在花名册内的记录）"
            flash(msg, "success")
            return redirect(url_for("performance.records", year=selected_year, month=selected_month))
        except Exception as exc:
            error_msg = f"解析失败：{str(exc)}"
            flash(error_msg, "danger")

            # 记录失败的导入操作
            log_import_operation(
                module='performance',
                operation='import',
                file_name=filename if 'filename' in locals() else None,
                total_rows=0,
                success_rows=0,
                failed_rows=0,
                skipped_rows=0,
                error_message=error_msg
            )

            return redirect(url_for("performance.upload"))

    return render_template(
        "upload.html",
        title=f"上传PDF | {APP_TITLE}",
        year_options=year_options,
        selected_year=selected_year,
        selected_month=selected_month,
        confirm_ctx=None,
    )


@performance_bp.route('/records')
@login_required
def records():
    """绩效记录列表（单月或年度视图）"""
    from flask import session
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    sort = request.args.get("sort", default="avg_desc")
    selected_grades = [g for g in request.args.getlist("grade") if g]
    grade_filter_set = set(selected_grades)
    min_score_raw = (request.args.get("min_score") or "").strip()
    try:
        min_score_filter = float(min_score_raw) if min_score_raw else None
    except ValueError:
        min_score_filter = None
        min_score_raw = ""

    uid = require_user_id()
    now = datetime.now()
    year_options = list(range(now.year - 5, now.year + 2))
    if year and year not in year_options:
        year_options.append(year)
        year_options.sort()

    # 获取当前用户角色
    conn = get_db()
    cur = conn.cursor()
    user_id = session.get('user_id')
    cur.execute("SELECT role FROM users WHERE id = ?", (user_id,))
    row = cur.fetchone()
    user_role = row['role'] if row else 'user'

    if month and year:
        # 使用部门过滤机制
        where_clause, join_clause, dept_params = build_department_filter('pr')
        cur.execute(
            f"""
            SELECT pr.id, pr.emp_no, pr.name, pr.year, pr.month, pr.score, pr.grade
            FROM performance_records pr
            {join_clause}
            WHERE {where_clause} AND pr.year=? AND pr.month=?
            ORDER BY CAST(pr.emp_no as INTEGER)
            """,
            dept_params + [year, month],
        )
        rows = cur.fetchall()
        period_title = f"{year}年{month}月"
        avg_score = None
        if rows:
            scores = [r["score"] for r in rows if r["score"] is not None]
            if scores:
                avg_score = round(sum(scores) / len(scores), 2)
        grade_choices = sorted({r["grade"] for r in rows if r["grade"]})
        grade_select_size = min(max(len(grade_choices), 4), 8) if grade_choices else 4
        return render_template(
            "records.html",
            title=f"记录 | {APP_TITLE}",
            view_mode="single",
            rows=rows,
            period_title=period_title,
            avg_score=avg_score,
            year=year,
            month=month,
            sort=sort,
            grade_choices=grade_choices,
            selected_grades=selected_grades,
            min_score_filter=min_score_raw,
            grade_select_size=grade_select_size,
            year_options=year_options,
            user_role=user_role,
        )

    year = year or now.year
    if year not in year_options:
        year_options.append(year)
        year_options.sort()

    data, months, grade_choices = build_yearly_matrix(uid, year)
    data = filter_sort_yearly_data(data, grade_filter_set, min_score_filter, sort)
    grade_select_size = min(max(len(grade_choices), 4), 8) if grade_choices else 4

    return render_template(
        "records.html",
        title=f"记录 | {APP_TITLE}",
        view_mode="dashboard",
        data=data,
        months=months,
        year=year,
        month=None,
        sort=sort,
        period_title=f"{year}年（1-12月）",
        grade_choices=grade_choices,
        selected_grades=selected_grades,
        min_score_filter=min_score_raw,
        grade_select_size=grade_select_size,
        year_options=year_options,
        user_role=user_role,
    )


@performance_bp.route('/range')
@login_required
def range_view():
    """自定义区间绩效统计"""
    # 支持新格式（YYYY-MM）和旧格式（sy, sm, ey, em）的向后兼容
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    # 如果使用新格式，解析为年月
    if start_date and end_date:
        try:
            sy, sm = map(int, start_date.split('-'))
            ey, em = map(int, end_date.split('-'))
        except (ValueError, AttributeError):
            sy = sm = ey = em = None
    else:
        # 向后兼容：使用旧的参数格式
        sy = request.args.get("sy", type=int)
        sm = request.args.get("sm", type=int)
        ey = request.args.get("ey", type=int)
        em = request.args.get("em", type=int)
        # 如果有旧参数，转换为新格式供模板使用
        if sy and sm:
            start_date = f"{sy:04d}-{sm:02d}"
        if ey and em:
            end_date = f"{ey:04d}-{em:02d}"

    data = []
    months = []
    uid = require_user_id()

    if sy and sm and ey and em:
        start = ym_to_int(sy, sm)
        end = ym_to_int(ey, em)
        if start <= end:
            y, m = sy, sm
            while True:
                months.append((y, m))
                if y == ey and m == em:
                    break
                m += 1
                if m == 13:
                    m = 1
                    y += 1

            conn = get_db()
            cur = conn.cursor()

            # 使用部门过滤机制
            where_clause, join_clause, dept_params = build_department_filter('pr')
            cur.execute(
                f"""
                SELECT pr.emp_no, pr.name, pr.year, pr.month, pr.score, pr.grade
                FROM performance_records pr
                {join_clause}
                WHERE {where_clause} AND (pr.year*100 + pr.month) BETWEEN ? AND ?
                ORDER BY CAST(pr.emp_no as INTEGER), pr.year, pr.month
                """,
                dept_params + [start, end],
            )
            rows = cur.fetchall()

            by_emp = {}
            for r in rows:
                key = r["emp_no"]
                by_emp.setdefault(key, {"emp_no": r["emp_no"], "name": r["name"], "items": {}})
                by_emp[key]["items"][(r["year"], r["month"])] = {
                    "score": r["score"],
                    "grade": r["grade"],
                }

            data = []
            for emp in by_emp.values():
                row = {"emp_no": emp["emp_no"], "name": emp["name"], "detail": []}
                scores = []
                for (y, m) in months:
                    item = emp["items"].get((y, m))
                    if item:
                        row["detail"].append(item)
                        if item["score"] is not None:
                            scores.append(item["score"])
                    else:
                        row["detail"].append(None)
                row["avg_score"] = round(sum(scores) / len(scores), 2) if scores else None
                data.append(row)
            data.sort(key=lambda r: emp_sort_key(r["emp_no"]))

    return render_template(
        "range.html",
        title=f"区间统计 | {APP_TITLE}",
        data=data,
        months=months,
        start_date=start_date,
        end_date=end_date,
        sy=sy,  # 保留旧参数以支持向后兼容
        sm=sm,
        ey=ey,
        em=em,
    )


@performance_bp.route('/export')
@login_required
def export_single():
    """导出单月绩效数据"""
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        flash("请在列表页面选择具体的 年/月 后再导出", "warning")
        return redirect(url_for("performance.records"))

    uid = require_user_id()
    conn = get_db()
    cur = conn.cursor()

    # 使用部门过滤机制
    where_clause, join_clause, dept_params = build_department_filter('pr')
    cur.execute(
        f"""
        SELECT pr.emp_no, pr.name, pr.year, pr.month, pr.score, pr.grade
        FROM performance_records pr
        {join_clause}
        WHERE {where_clause} AND pr.year=? AND pr.month=?
        ORDER BY CAST(pr.emp_no as INTEGER)
        """,
        dept_params + [year, month],
    )
    rows = cur.fetchall()

    if not rows:
        flash("无数据可导出", "warning")
        return redirect(url_for("performance.records", year=year, month=month))

    xlsx_path = os.path.join(EXPORT_DIR, f"绩效_{year}-{str(month).zfill(2)}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{str(month).zfill(2)}"
    ws.append(["工号", "姓名", "年份", "月份", "分数", "等级"])
    for r in rows:
        ws.append([r["emp_no"], r["name"], r["year"], r["month"], r["score"], r["grade"]])
    wb.save(xlsx_path)
    return send_file(xlsx_path, as_attachment=True, download_name=os.path.basename(xlsx_path))


@performance_bp.route('/export_range')
@login_required
def export_range():
    """导出区间绩效数据"""
    # 支持新格式（YYYY-MM）和旧格式（sy, sm, ey, em）的向后兼容
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    # 如果使用新格式，解析为年月
    if start_date and end_date:
        try:
            sy, sm = map(int, start_date.split('-'))
            ey, em = map(int, end_date.split('-'))
        except (ValueError, AttributeError):
            sy = sm = ey = em = None
    else:
        # 向后兼容：使用旧的参数格式
        sy = request.args.get("sy", type=int)
        sm = request.args.get("sm", type=int)
        ey = request.args.get("ey", type=int)
        em = request.args.get("em", type=int)

    if not (sy and sm and ey and em):
        flash("请先选择起止年月", "warning")
        return redirect(url_for("performance.range_view"))

    start = ym_to_int(sy, sm)
    end = ym_to_int(ey, em)
    if start > end:
        flash("起止月份不合法：起始不能大于结束", "warning")
        return redirect(url_for("performance.range_view", sy=sy, sm=sm, ey=ey, em=em))

    months = []
    y, m = sy, sm
    while True:
        months.append((y, m))
        if y == ey and m == em:
            break
        m += 1
        if m == 13:
            m = 1
            y += 1

    uid = require_user_id()
    conn = get_db()
    cur = conn.cursor()

    # 使用部门过滤机制
    where_clause, join_clause, dept_params = build_department_filter('pr')
    cur.execute(
        f"""
        SELECT pr.emp_no, pr.name, pr.year, pr.month, pr.score, pr.grade
        FROM performance_records pr
        {join_clause}
        WHERE {where_clause} AND (pr.year*100 + pr.month) BETWEEN ? AND ?
        ORDER BY CAST(pr.emp_no as INTEGER), pr.year, pr.month
        """,
        dept_params + [start, end],
    )
    rows = cur.fetchall()

    by_emp = {}
    for r in rows:
        key = r["emp_no"]
        by_emp.setdefault(key, {"emp_no": r["emp_no"], "name": r["name"], "items": {}})
        by_emp[key]["items"][(r["year"], r["month"])] = {
            "score": r["score"],
            "grade": r["grade"],
        }

    xlsx_path = os.path.join(
        EXPORT_DIR,
        f"绩效_区间_{sy}-{str(sm).zfill(2)}_到_{ey}-{str(em).zfill(2)}.xlsx",
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "区间统计"

    headers = ["工号", "姓名"]
    for (y, m) in months:
        ym = f"{y}-{str(m).zfill(2)}"
        headers.extend([f"{ym}分数", f"{ym}等级"])
    headers.append("区间平均分")
    ws.append(headers)

    def avg_score(vals):
        vals = [float(x) for x in vals if x is not None]
        return round(sum(vals) / len(vals), 2) if vals else ""

    for emp in sorted(by_emp.values(), key=lambda e: emp_sort_key(e["emp_no"])):
        line = [emp["emp_no"], emp["name"]]
        scores_acc = []
        for (y, m) in months:
            item = emp["items"].get((y, m))
            score = item["score"] if item else None
            grade = item["grade"] if item else ""
            line.append(score if score is not None else "")
            line.append(grade)
            if score is not None:
                scores_acc.append(score)
        line.append(avg_score(scores_acc))
        ws.append(line)

    wb.save(xlsx_path)
    return send_file(xlsx_path, as_attachment=True, download_name=os.path.basename(xlsx_path))


@performance_bp.route('/export_yearly')
@login_required
def export_yearly():
    """导出年度绩效数据"""
    uid = require_user_id()
    year = request.args.get("year", type=int)
    if not year:
        flash("请先选择年份", "warning")
        return redirect(url_for("performance.records"))

    sort = request.args.get("sort", "avg_desc")
    selected_grades = [g for g in request.args.getlist("grade") if g]
    grade_filter_set = set(selected_grades)
    min_score_raw = (request.args.get("min_score") or "").strip()
    try:
        min_score_filter = float(min_score_raw) if min_score_raw else None
    except ValueError:
        min_score_filter = None

    data, months, _ = build_yearly_matrix(uid, year)
    data = filter_sort_yearly_data(data, grade_filter_set, min_score_filter, sort)
    if not data:
        flash("无数据可导出", "warning")
        return redirect(url_for("performance.records", year=year, sort=sort))

    xlsx_path = os.path.join(EXPORT_DIR, f"绩效年度_{year}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = f"年度{year}"

    headers = ["工号", "姓名"]
    for _y, m in months:
        headers.extend([f"{m}月分数", f"{m}月等级"])
    headers.append("平均分")
    ws.append(headers)

    for row in data:
        line = [row["emp_no"], row["name"]]
        for cell in row["detail"]:
            if cell:
                line.append(cell.get("score") if cell.get("score") is not None else "")
                line.append(cell.get("grade", ""))
            else:
                line.extend(["", ""])
        line.append(row["avg_score"] if row["avg_score"] is not None else "")
        ws.append(line)

    wb.save(xlsx_path)
    return send_file(
        xlsx_path,
        as_attachment=True,
        download_name=os.path.basename(xlsx_path),
    )


@performance_bp.route('/export_quarters')
@login_required
def export_quarters():
    """导出季度绩效数据"""
    uid = require_user_id()
    year = request.args.get("year", type=int) or datetime.now().year
    selected_grades = [g for g in request.args.getlist("grade") if g]
    grade_filter_set = set(selected_grades)
    min_score_raw = (request.args.get("min_score") or "").strip()
    try:
        min_score_filter = float(min_score_raw) if min_score_raw else None
    except ValueError:
        min_score_filter = None

    options, default_grade, data = build_quarter_dataset(uid, year)
    data = filter_quarter_data(data, grade_filter_set, min_score_filter)
    if not data:
        flash("无数据可导出", "warning")
        return redirect(url_for("performance.quarters", year=year))

    xlsx_path = os.path.join(EXPORT_DIR, f"季度绩效_{year}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = f"季度{year}"

    headers = [
        "工号",
        "姓名",
        "Q1平均分",
        "Q1等级",
        "Q2平均分",
        "Q2等级",
        "Q3平均分",
        "Q3等级",
        "Q4平均分",
        "Q4等级",
    ]
    ws.append(headers)

    for row in data:
        line = [row["emp_no"], row["name"]]
        for quarter in (1, 2, 3, 4):
            cell = row["q"].get(quarter, {})
            score = cell.get("score")
            line.append(score if isinstance(score, (int, float)) else (score or ""))
            line.append(cell.get("grade") or default_grade)
        ws.append(line)

    wb.save(xlsx_path)
    return send_file(
        xlsx_path,
        as_attachment=True,
        download_name=os.path.basename(xlsx_path),
    )


@performance_bp.route('/export_calculator')
@login_required
def export_calculator():
    """导出绩效计算器结果"""
    uid = require_user_id()
    sort = request.args.get("sort", "total_desc")
    year = request.args.get("year", type=int) or datetime.now().year

    mapping = get_or_init_grade_map(uid)
    data, months = build_calculator_dataset(uid, year, mapping)
    data = sort_calculator_data(data, sort)
    if not data:
        flash("无数据可导出", "warning")
        return redirect(url_for("performance.calculator", sort=sort, year=year))

    xlsx_path = os.path.join(EXPORT_DIR, f"绩效计算器_{year}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = f"积分{year}"

    headers = ["工号", "姓名"]
    for _y, m in months:
        headers.extend([f"{m}月等级", f"{m}月积分"])
    headers.append("总积分")
    ws.append(headers)

    for row in data:
        line = [row["emp_no"], row["name"]]
        for cell in row["detail"]:
            if cell:
                line.append(cell.get("grade", ""))
                line.append(round(cell.get("value", 0.0), 2))
            else:
                line.extend(["", ""])
        line.append(row["total"])
        ws.append(line)

    wb.save(xlsx_path)
    return send_file(
        xlsx_path,
        as_attachment=True,
        download_name=os.path.basename(xlsx_path),
    )


@performance_bp.route('/calculator', methods=['GET', 'POST'])
@login_required
def calculator():
    """绩效计算器（自定义档位映射）"""
    uid = require_user_id()
    sort = request.args.get("sort", "total_desc")
    year_param = request.args.get("year", type=int)
    now = datetime.now()
    year = year_param or now.year
    if not (2000 <= year <= 2100):
        year = now.year
    year_options = list(range(now.year - 5, now.year + 2))
    if year not in year_options:
        year_options.append(year)
        year_options.sort()

    mapping = get_or_init_grade_map(uid)

    if request.method == "POST":
        sort = request.form.get("sort", sort) or "total_desc"
        year_form = (request.form.get("year") or "").strip()
        if year_form:
            try:
                year = int(year_form)
            except ValueError:
                pass
        if not (2000 <= year <= 2100):
            year = now.year
        if year not in year_options:
            year_options.append(year)
            year_options.sort()

        def _num(value):
            try:
                return float(value)
            except Exception:
                return 0.0

        new_map = {
            "A": _num(request.form.get("score_A", mapping.get("A", 5))),
            "B+": _num(request.form.get("score_Bp", mapping.get("B+", 4))),
            "B": _num(request.form.get("score_B", mapping.get("B", 3))),
            "C": _num(request.form.get("score_C", mapping.get("C", 2))),
            "D": _num(request.form.get("score_D", mapping.get("D", 1))),
            "不进行": _num(request.form.get("score_none", mapping.get("不进行", 0))),
        }
        conn = get_db()
        cur = conn.cursor()
        for grade, value in new_map.items():
            cur.execute(
                "INSERT OR REPLACE INTO grade_map(grade, value) VALUES(?,?)",
                (grade, value),
            )
        conn.commit()
        flash("分值设置已保存。", "success")
        return redirect(url_for("performance.calculator", sort=sort, year=year))

    data, months = build_calculator_dataset(uid, year, mapping)
    data = sort_calculator_data(data, sort)

    return render_template(
        "calculator.html",
        title=f"绩效计算器 | {APP_TITLE}",
        mapping=mapping,
        data=data,
        months=months,
        year=year,
        sort=sort,
        year_options=year_options,
    )


@performance_bp.route('/quarters', methods=['GET', 'POST'])
@login_required
def quarters():
    """季度绩效管理（含人工修正）"""
    uid = require_user_id()
    year = request.args.get("year", type=int) or datetime.now().year
    selected_grades = [g for g in request.args.getlist("grade") if g]
    grade_filter_set = set(selected_grades)
    min_score_raw = (request.args.get("min_score") or "").strip()
    try:
        min_score_filter = float(min_score_raw) if min_score_raw else None
    except ValueError:
        min_score_filter = None
        min_score_raw = ""
    options = get_quarter_grade_options(uid)
    default_grade = next((opt["grade"] for opt in options if opt["is_default"]), options[0]["grade"])
    grade_select_size = min(max(len(options), 4), 8) if options else 4

    if request.method == "POST":
        form_kind = request.form.get("form")
        if form_kind == "options":
            raw = (request.form.get("grades_input", "") or "").strip()
            entries: List[Tuple[str, str]] = []
            seen = set()
            for raw_line in raw.splitlines():
                line = raw_line.strip()
                if not line:
                    continue
                parts = [p.strip() for p in line.replace("，", ",").split(",") if p.strip()]
                if not parts:
                    continue
                grade = parts[0]
                if grade in seen:
                    continue
                base_color = DEFAULT_QUARTER_COLORS.get(grade, "#FFFFFF")
                color = parts[1] if len(parts) > 1 else base_color
                color = normalize_color(color, base_color)
                seen.add(grade)
                entries.append((grade, color))
            if not entries:
                flash("请至少填写一个季度等级选项。", "warning")
                return redirect(url_for("performance.quarters", year=year))
            chosen_default = request.form.get("default_grade", "").strip()
            grade_list = [grade for grade, _ in entries]
            if chosen_default not in grade_list:
                chosen_default = grade_list[0]
            set_quarter_grade_options(uid, entries, chosen_default)
            flash("季度等级选项已更新。", "success")
            return redirect(url_for("performance.quarters", year=year))

        conn = get_db()
        cur = conn.cursor()
        for key, value in request.form.items():
            if not key.startswith("grade_q"):
                continue
            try:
                _, tail = key.split("q", 1)
                q_str, emp_no = tail.split("_", 1)
                quarter = int(q_str)
            except Exception:
                continue
            grade = (value or default_grade).strip()
            cur.execute(
                """
                INSERT INTO quarter_overrides(emp_no, year, quarter, grade)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(emp_no, year, quarter) DO UPDATE SET grade=excluded.grade
                """,
                (emp_no, year, quarter, grade),
            )
        conn.commit()
        flash("季度绩效等级已保存。", "success")
        return redirect(url_for("performance.quarters", year=year))

    options, default_grade, data = build_quarter_dataset(uid, year, options)
    data = filter_quarter_data(data, grade_filter_set, min_score_filter)
    grade_select_size = min(max(len(options), 4), 8) if options else 4

    return render_template(
        "quarters.html",
        title=f"季度绩效 | {APP_TITLE}",
        data=data,
        year=year,
        grade_options=options,
        grade_default=default_grade,
        selected_grades=selected_grades,
        min_score_filter=min_score_raw,
        grade_select_size=grade_select_size,
    )


@performance_bp.route('/records/<int:record_id>/edit', methods=['POST'])
@role_required('manager')
def edit_record(record_id):
    """编辑绩效记录（仅限部门管理员及以上权限）"""
    conn = get_db()
    cur = conn.cursor()

    # 获取表单数据
    emp_no = request.form.get('emp_no', '').strip()
    name = request.form.get('name', '').strip()
    year = request.form.get('year', '').strip()
    month = request.form.get('month', '').strip()
    score = request.form.get('score', '').strip()
    grade = request.form.get('grade', '').strip()

    # 验证必填字段
    if not emp_no or not name or not year or not month or not score or not grade:
        flash('所有字段均为必填项', 'warning')
        return redirect(url_for('performance.records'))

    try:
        # 更新记录
        cur.execute("""
            UPDATE performance_records
            SET emp_no = ?, name = ?, year = ?, month = ?, score = ?, grade = ?
            WHERE id = ?
        """, (emp_no, name, int(year), int(month), float(score), grade, record_id))

        conn.commit()
        flash('绩效记录已更新', 'success')
    except Exception as e:
        flash(f'更新失败: {e}', 'danger')

    return redirect(url_for('performance.records'))


@performance_bp.route('/records/<int:record_id>/delete', methods=['POST'])
@role_required('manager')
def delete_record(record_id):
    """删除绩效记录（仅限部门管理员及以上权限）"""
    conn = get_db()
    cur = conn.cursor()

    try:
        # 删除记录
        cur.execute("DELETE FROM performance_records WHERE id = ?", (record_id,))
        conn.commit()
        flash('绩效记录已删除', 'success')
    except Exception as e:
        flash(f'删除失败: {e}', 'danger')

    return redirect(url_for('performance.records'))


@performance_bp.route('/records/batch-delete', methods=['POST'])
@role_required('manager')
def batch_delete_records():
    """批量删除绩效记录（仅限部门管理员及以上权限）"""
    conn = get_db()
    cur = conn.cursor()

    record_ids = request.form.getlist('record_ids')

    if not record_ids:
        flash('未选择要删除的记录', 'warning')
        return redirect(url_for('performance.records'))

    try:
        # 批量删除记录
        placeholders = ','.join('?' * len(record_ids))
        cur.execute(f"DELETE FROM performance_records WHERE id IN ({placeholders})", record_ids)
        conn.commit()
        flash(f'成功删除 {len(record_ids)} 条绩效记录', 'success')
    except Exception as e:
        flash(f'批量删除失败: {e}', 'danger')

    return redirect(url_for('performance.records'))
