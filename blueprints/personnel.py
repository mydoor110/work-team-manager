#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
人员管理模块
负责员工信息管理、导入导出等功能
"""
import json
import sqlite3
from collections import Counter
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from openpyxl import Workbook, load_workbook

from config.settings import APP_TITLE
from models.database import get_db
from .decorators import login_required, manager_required
from .helpers import (
    current_user_id, require_user_id, get_accessible_department_ids,
    get_accessible_departments, calculate_years_from_date, get_user_department,
    validate_employee_access, log_import_operation
)

# 创建 Blueprint
personnel_bp = Blueprint('personnel', __name__, url_prefix='/personnel')


# ==================== 常量定义 ====================

PERSONNEL_FIELD_SCHEME = [
    {"name": "emp_no", "label": "工号", "input_type": "text", "required": True},
    {"name": "name", "label": "姓名", "input_type": "text", "required": True},
    {"name": "department_id", "label": "所属部门", "input_type": "department_select", "required": True},
    {"name": "class_name", "label": "班级", "input_type": "text"},
    {"name": "position", "label": "岗位", "input_type": "text"},
    {"name": "birth_date", "label": "出生年月", "input_type": "date"},
    {"name": "certification_date", "label": "取证时间", "input_type": "date"},
    {"name": "solo_driving_date", "label": "单独驾驶时间", "input_type": "date"},
    {"name": "marital_status", "label": "婚姻状况", "input_type": "select"},
    {"name": "hometown", "label": "籍贯", "input_type": "text"},
    {"name": "political_status", "label": "政治面貌", "input_type": "select"},
    {"name": "education", "label": "学历", "input_type": "select"},
    {"name": "graduation_school", "label": "毕业院校", "input_type": "text"},
    {"name": "work_start_date", "label": "参加工作时间", "input_type": "date"},
    {"name": "entry_date", "label": "入司时间", "input_type": "date"},
    {"name": "specialty", "label": "特长及兴趣爱好", "input_type": "textarea"},
]

PERSONNEL_DB_COLUMNS = [
    field["name"] for field in PERSONNEL_FIELD_SCHEME if field["name"] not in {"emp_no", "name"}
]

PERSONNEL_DATE_FIELDS = {"birth_date", "work_start_date", "entry_date", "certification_date", "solo_driving_date"}

PERSONNEL_SELECT_OPTIONS = {
    "marital_status": ["未婚", "已婚", "离异", "其它"],
    "political_status": ["中共党员", "中共预备党员", "共青团员", "群众", "其它"],
    "education": ["博士研究生", "硕士研究生", "本科", "大专", "中专", "高中", "其它"],
}

PERSONNEL_IMPORT_HEADER_MAP = {
    "工号": "emp_no",
    "姓名": "name",
    "所属部门": "department_id",
    "部门": "department_id",
    "班级": "class_name",
    "岗位": "position",
    "出生年月": "birth_date",
    "取证时间": "certification_date",
    "取证日期": "certification_date",
    "单独驾驶时间": "solo_driving_date",
    "单独驾驶日期": "solo_driving_date",
    "婚否": "marital_status",
    "婚姻状况": "marital_status",
    "籍贯": "hometown",
    "政治面貌": "political_status",
    "特长及兴趣爱好": "specialty",
    "特长": "specialty",
    "学历": "education",
    "毕业院校": "graduation_school",
    "参加工作时间": "work_start_date",
    "入司时间": "entry_date",
}


# ==================== 辅助函数 ====================

def calculate_performance_score_monthly(grade: str, raw_score: float, config: dict = None) -> Dict:
    """
    绩效月度快照算法（参数化版本）

    Args:
        grade: 绩效等级 (A, B+, B, C, D)
        raw_score: 原始计算分 (100 + 加分 - 扣分)
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'radar_value': 雷达图显示值,
            'display_label': 显示标签,
            'status_color': 状态颜色 (RED/ORANGE/GREEN),
            'alert_tag': 警示标签,
            'grade': 等级
        }
    """
    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    grade_coefficients = config['performance']['grade_coefficients']
    grade_ranges = config['performance']['grade_ranges']

    grade = grade.upper() if grade else 'B+'

    # 等级锁定规则（使用配置参数）
    if grade == 'D':
        radar_value = grade_ranges['D']['radar_override']  # 从配置读取
        status_color = 'RED'
        alert_tag = '⛔ 绩效不合格'
        display_label = f'D级 (系数{grade_coefficients["D"]})'
    elif grade == 'C':
        radar_value = min(max(raw_score, grade_ranges['C']['min']), grade_ranges['C']['max'])
        status_color = 'ORANGE'
        alert_tag = '⚠️ 绩效预警'
        display_label = f'C级 (系数{grade_coefficients["C"]})'
    elif grade == 'B':
        radar_value = min(max(raw_score, grade_ranges['B']['min']), grade_ranges['B']['max'])
        status_color = 'ORANGE'
        alert_tag = '⚠️ 未达基准'
        display_label = f'B级 (系数{grade_coefficients["B"]})'
    elif grade == 'B+':
        radar_value = min(max(raw_score, grade_ranges['B+']['min']), grade_ranges['B+']['max'])
        status_color = 'GREEN'
        alert_tag = '✅ 达标'
        display_label = f'B+级 (系数{grade_coefficients["B+"]})'
    elif grade == 'A':
        radar_value = min(max(raw_score, grade_ranges['A']['min']), grade_ranges['A']['max'])
        status_color = 'GREEN'
        alert_tag = '✅ 优秀'
        display_label = f'A级 (系数{grade_coefficients["A"]})'
    else:  # 默认B+
        radar_value = min(max(raw_score, grade_ranges['B+']['min']), grade_ranges['B+']['max'])
        status_color = 'GREEN'
        alert_tag = '✅ 达标'
        display_label = f'B+级 (系数{grade_coefficients["B+"]})'

    return {
        'radar_value': round(radar_value, 1),
        'display_label': display_label,
        'status_color': status_color,
        'alert_tag': alert_tag,
        'grade': grade,
        'mode': 'MONTHLY'
    }


def calculate_performance_score_period(grade_list: List[str], grade_dates: Optional[List[str]] = None, config: dict = None) -> Dict:
    """
    绩效周期加权算法（跨月、季度、年度）（参数化版本）

    新增时间衰减机制：D级和C级的影响会随时间推移而减弱

    Args:
        grade_list: 周期内所有月份的等级列表，如 ['A', 'B+', 'B', 'C']
        grade_dates: 每个等级对应的日期列表（可选），如 ['2024-01', '2024-02', ...]
                     如果提供，将启用时间衰减机制
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'radar_value': 雷达图显示值,
            'display_label': 显示标签,
            'status_color': 状态颜色,
            'alert_tag': 警示标签
        }
    """
    if not grade_list:
        return {
            'radar_value': 95.0,
            'display_label': '暂无数据',
            'status_color': 'GREEN',
            'alert_tag': '✅ 暂无数据',
            'mode': 'PERIOD'
        }

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    grade_coefficients = config['performance']['grade_coefficients']
    contamination_rules = config['performance']['contamination_rules']
    time_decay = config['performance'].get('time_decay', {
        'enabled': True,
        'decay_months': 6,
        'decay_rate': 0.9
    })

    # Step 1: 系数映射（使用配置）
    coeff_map = grade_coefficients

    coeffs = []
    d_count = 0
    c_count = 0
    d_count_effective = 0.0  # 带时间衰减的有效D级计数
    c_count_effective = 0.0  # 带时间衰减的有效C级计数

    # 如果启用时间衰减且提供了日期信息
    use_time_decay = time_decay.get('enabled', True) and grade_dates and len(grade_dates) == len(grade_list)

    if use_time_decay:
        from datetime import datetime

        now = datetime.now()
        decay_months_threshold = time_decay.get('decay_months', 6)
        decay_rate_per_month = time_decay.get('decay_rate', 0.9)

        for i, (grade, date_str) in enumerate(zip(grade_list, grade_dates)):
            grade = grade.upper() if grade else 'B+'
            coeffs.append(coeff_map.get(grade, 1.0))

            try:
                # 解析日期（支持 YYYY-MM 或 YYYY-MM-DD 格式）
                if len(date_str) == 7:  # YYYY-MM
                    grade_date = datetime.strptime(date_str, '%Y-%m')
                else:  # YYYY-MM-DD
                    grade_date = datetime.strptime(date_str[:7], '%Y-%m')

                # 计算距今月数
                months_ago = (now.year - grade_date.year) * 12 + (now.month - grade_date.month)

                if grade == 'D':
                    d_count += 1
                    # 时间衰减逻辑：
                    # 1. 只计入最近 decay_months_threshold 个月内的D级
                    # 2. 每个月衰减 (1 - decay_rate_per_month)
                    if months_ago <= decay_months_threshold:
                        # 计算衰减权重：decay_rate^months_ago
                        decay_weight = (decay_rate_per_month ** months_ago)
                        d_count_effective += decay_weight
                elif grade == 'C':
                    c_count += 1
                    # C级同样应用时间衰减（但阈值和惩罚可能不同）
                    if months_ago <= decay_months_threshold:
                        decay_weight = (decay_rate_per_month ** months_ago)
                        c_count_effective += decay_weight

            except Exception:
                # 日期解析失败，按原逻辑计数
                if grade == 'D':
                    d_count += 1
                    d_count_effective += 1
                elif grade == 'C':
                    c_count += 1
                    c_count_effective += 1
    else:
        # 不使用时间衰减，按原逻辑
        for grade in grade_list:
            grade = grade.upper() if grade else 'B+'
            coeffs.append(coeff_map.get(grade, 1.0))
            if grade == 'D':
                d_count += 1
                d_count_effective = d_count
            elif grade == 'C':
                c_count += 1
                c_count_effective = c_count

    # Step 2: 计算平均系数
    avg_coeff = sum(coeffs) / len(coeffs) if coeffs else 1.0

    # Step 3: 还原基础分 (系数1.0对应95分)
    base_score = avg_coeff * 95

    # Step 4: 执行"污点熔断"规则（使用时间衰减后的计数）
    d_threshold = contamination_rules['d_count_threshold']
    c_threshold = contamination_rules['c_count_threshold']
    d_cap = contamination_rules['d_cap_score']
    c_cap = contamination_rules['c_cap_score']

    if d_count_effective >= d_threshold:
        # D级熔断规则（使用衰减后的计数）
        final_score = min(base_score, d_cap)
        status_color = 'RED'
        if use_time_decay and d_count_effective < d_count:
            alert_tag = f'⛔ 存在D级考核 (有效{d_count_effective:.1f}次)'
        else:
            alert_tag = '⛔ 存在D级考核'
    elif c_count_effective >= c_threshold:
        # C级熔断规则（使用衰减后的计数）
        final_score = min(base_score, c_cap)
        status_color = 'ORANGE'
        if use_time_decay and c_count_effective < c_count:
            alert_tag = f'⚠️ 多次C级预警 (有效{c_count_effective:.1f}次)'
        else:
            alert_tag = '⚠️ 多次C级预警'
    else:
        # 正常输出
        final_score = min(base_score, 110)
        if final_score >= 95:
            status_color = 'GREEN'
            alert_tag = '✅ 综合达标'
        elif final_score >= 80:
            status_color = 'ORANGE'
            alert_tag = '⚠️ 未达基准'
        else:
            status_color = 'RED'
            alert_tag = '⛔ 综合不合格'

    # 生成显示标签
    display_label = f'平均系数{avg_coeff:.2f}'

    return {
        'radar_value': round(final_score, 1),
        'display_label': display_label,
        'status_color': status_color,
        'alert_tag': alert_tag,
        'mode': 'PERIOD',
        'd_count_raw': d_count,  # 原始D级次数
        'd_count_effective': round(d_count_effective, 2),  # 时间衰减后有效次数
        'time_decay_applied': use_time_decay
    }


def calculate_safety_score_dual_track(violations_list: List[float], months_active: int = 1, config: dict = None) -> Dict:
    """
    安全意识双轨评分模型（参数化版本）

    Args:
        violations_list: 违规扣分值列表，例如 [1, 3, 6]
        months_active: 统计周期包含的月份数（月度传1，年度传12或实际在职月数）
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'score_a': 行为分（习惯维度）,
            'score_b': 严重性分（后果维度）,
            'final_score': 最终分数（取两者最低）,
            'status_color': 状态颜色（RED/ORANGE/GREEN）,
            'alert_tag': 警示标签
        }
    """
    import math

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    behavior_track = config['safety']['behavior_track']
    severity_track = config['safety']['severity_track']
    thresholds = config['safety']['thresholds']

    # 维度A：行为习惯（捉拿惯犯）
    violation_count = len(violations_list)
    avg_freq = math.ceil(violation_count / months_active) if months_active > 0 else 0

    # 根据月均频次扣分（使用配置参数）
    freq_thresholds = behavior_track['freq_thresholds']  # [2, 5, 6]
    freq_multipliers = behavior_track['freq_multipliers']  # [2, 5, 10]

    if avg_freq <= freq_thresholds[0]:
        score_a_deduction = avg_freq * freq_multipliers[0]
    elif freq_thresholds[0] < avg_freq <= freq_thresholds[1]:
        score_a_deduction = avg_freq * freq_multipliers[1]
    else:  # avg_freq >= freq_thresholds[2]
        score_a_deduction = avg_freq * freq_multipliers[2]

    score_a = max(0, 100 - score_a_deduction)

    # 维度B：后果严重性（精准打击）（使用配置参数）
    score_b_deduction = 0
    critical_threshold = severity_track['critical_threshold']
    has_critical_violation = False

    for score_value in violations_list:
        # 根据配置的score_ranges确定系数
        multiplier = 1.0
        for range_rule in severity_track['score_ranges']:
            if 'max' in range_rule and 'min' not in range_rule:
                # 只有max，表示 < max
                if score_value < range_rule['max']:
                    multiplier = range_rule['multiplier']
                    break
            elif 'min' in range_rule and 'max' in range_rule:
                # 有min和max，表示范围
                if range_rule['min'] <= score_value < range_rule['max']:
                    multiplier = range_rule['multiplier']
                    break
            elif 'min' in range_rule and 'max' not in range_rule:
                # 只有min，表示 >= min
                if score_value >= range_rule['min']:
                    multiplier = range_rule['multiplier']
                    break

        score_b_deduction += score_value * multiplier

        if score_value >= critical_threshold:
            has_critical_violation = True

    score_b = max(0, 100 - score_b_deduction)

    # 最终分数：取两者最低
    final_score = min(score_a, score_b)

    # 警示逻辑（使用配置阈值）
    fail_score = thresholds['fail_score']
    warning_score = thresholds['warning_score']

    if final_score < fail_score or has_critical_violation:
        # 红线熔断
        status_color = "RED"
        alert_tag = "⛔ 重大红线（存在高扣分）" if has_critical_violation else "⛔ 安全不合格"
    elif fail_score <= final_score < warning_score:
        # 黄色预警
        status_color = "ORANGE"
        if score_a < score_b:
            alert_tag = "⚠️ 高频违规风险"
        else:
            alert_tag = "⚠️ 扣分过多风险"
    else:  # final_score >= warning_score
        # 绿色安全
        status_color = "GREEN"
        alert_tag = "✅ 安全"

    return {
        'score_a': round(score_a, 1),
        'score_b': round(score_b, 1),
        'final_score': round(final_score, 1),
        'status_color': status_color,
        'alert_tag': alert_tag,
        'violation_count': violation_count,
        'avg_freq': avg_freq
    }


def calculate_training_score_with_penalty(
    training_records: List[tuple],
    duration_days: int = 30,
    cert_years: Optional[float] = None,
    config: dict = None
) -> Dict:
    """
    培训/实操能力高级评分算法 - 包含毒性惩罚和动态年化（参数化版本）

    新增动态AFR阈值：根据取证年限区分新老员工，使用不同的评判标准

    Args:
        training_records: 培训记录列表，每条记录为 (score, is_qualified, is_disqualified, training_date)
        duration_days: 统计周期天数（用于年化计算）
        cert_years: 取证年限（可选），用于判断新老员工。
                    None 或 <1年 为新员工，>=1年为老员工
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        dict: {
            'radar_score': 最终雷达图分数（已惩罚）,
            'original_score': 原始基础分,
            'penalty_coefficient': 惩罚系数,
            'stats': {'total_ops', 'fail_count', 'duration_days'},
            'risk_alert': {'show', 'level', 'text', 'description'},
            'status_color': 状态颜色（用于前端显示）
        }
    """
    import math

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    penalty_rules = config['training']['penalty_rules']
    duration_thresholds = config['training']['duration_thresholds']

    # Step 0: 数据准备
    total_ops = len(training_records)

    # 如果没有记录，根据统计周期判断严重程度（使用配置参数）
    if total_ops == 0:
        short_term_days = duration_thresholds['short_term_days']
        mid_term_days = duration_thresholds['mid_term_days']
        default_scores = duration_thresholds['default_scores']

        # 短期未培训：正常情况，给基础分
        if duration_days <= short_term_days:
            return {
                'radar_score': default_scores['short'],
                'original_score': default_scores['short'],
                'penalty_coefficient': 1.0,
                'stats': {
                    'total_ops': 0,
                    'fail_count': 0,
                    'duration_days': duration_days
                },
                'risk_alert': {
                    'show': True,
                    'level': 'NORMAL',
                    'text': '未开展培训',
                    'description': f'统计周期{duration_days}天内未开展培训，属于正常情况。'
                },
                'status_color': 'GREEN',
                'alert_tag': '未开展培训'
            }
        # 中期缺训：需要关注
        elif duration_days <= mid_term_days:
            return {
                'radar_score': default_scores['mid'],
                'original_score': default_scores['mid'],
                'penalty_coefficient': 1.0,
                'stats': {
                    'total_ops': 0,
                    'fail_count': 0,
                    'duration_days': duration_days
                },
                'risk_alert': {
                    'show': True,
                    'level': 'NOTICE',
                    'text': '⚠️ 长期未培训',
                    'description': f'统计周期{duration_days}天内未开展培训，建议安排培训。'
                },
                'status_color': 'YELLOW',
                'alert_tag': '⚠️ 长期未培训'
            }
        # 长期严重缺训：严重问题
        else:
            return {
                'radar_score': default_scores['long'],
                'original_score': default_scores['long'],
                'penalty_coefficient': 1.0,
                'stats': {
                    'total_ops': 0,
                    'fail_count': 0,
                    'duration_days': duration_days
                },
                'risk_alert': {
                    'show': True,
                    'level': 'CRITICAL',
                    'text': '❌ 严重缺训',
                    'description': f'统计周期{duration_days}天（超过半年）内未开展任何培训，严重影响业务能力。'
                },
                'status_color': 'RED',
                'alert_tag': '❌ 严重缺训'
            }

    # Step 1: 判定失格次数
    fail_count = 0
    total_score = 0

    for record in training_records:
        score, is_qualified, is_disqualified, training_date = record

        # 失格判定：is_disqualified=1 OR score=0 OR is_qualified=0
        if is_disqualified == 1 or score == 0 or is_qualified == 0:
            fail_count += 1

        total_score += (score if score else 0)

    # Step 2: 计算基础分（简单平均）
    avg_score = total_score / total_ops if total_ops > 0 else 0
    base_score = avg_score  # 可以根据需要调整权重，这里简化为平均分

    # Step 3: 确定惩罚系数（核心风控逻辑）
    coeff = 1.0
    tag_level = 'NORMAL'
    alert_msg = '✅ 能力达标'
    description = ''

    # Priority A: 绝对熔断红线（使用配置参数）
    absolute_threshold = penalty_rules['absolute_threshold']
    small_sample = penalty_rules['small_sample']

    if fail_count >= absolute_threshold['fail_count']:
        coeff = absolute_threshold['coefficient']
        tag_level = 'CRITICAL'
        alert_msg = '❌ 业务能力差 (高频失格)'
        description = f'检测到绝对失格次数 ≥ {absolute_threshold["fail_count"]}次（实际{fail_count}次），系统判定为不合格。'

    # Priority B: 小样本保护 & 高危标记（使用配置参数）
    elif total_ops < small_sample['sample_size'] and fail_count > 0:
        coeff = small_sample['coefficient']
        tag_level = 'HIGH_RISK'
        alert_msg = '⚠️ 观察期失格 (高风险-需带教)'
        description = f'样本量不足（仅{total_ops}次操作），但已出现{fail_count}次失格。建议加强带教。'

    # Priority C: 大样本年化推演（使用动态AFR阈值）
    elif total_ops >= small_sample['sample_size']:
        # 计算年化失格频率 (AFR - Annualized Failure Rate)
        duration_days = max(1, duration_days)  # 防止除零
        AFR = (fail_count / duration_days) * 365

        # 根据取证年限选择合适的AFR阈值（新增动态阈值逻辑）
        is_new_employee = cert_years is None or cert_years < 1.0

        if is_new_employee:
            # 新员工（取证1年内）：使用更宽松的阈值
            afr_thresholds = penalty_rules.get('afr_thresholds_new_employee', penalty_rules.get('afr_thresholds', []))
            employee_type = "新员工"
        else:
            # 老员工（取证1年以上）：使用标准阈值
            afr_thresholds = penalty_rules.get('afr_thresholds_experienced', penalty_rules.get('afr_thresholds', []))
            employee_type = "老员工"

        # 从高到低检查AFR阈值
        matched = False
        for rule in afr_thresholds:
            if 'max' in rule:
                # 有max的规则（中间范围）
                if rule['min'] <= AFR < rule['max']:
                    coeff = rule['coefficient']
                    tag_level = 'WARNING' if coeff <= 0.7 else 'NOTICE'
                    alert_msg = f'⛔ {rule["label"]} (年化 {AFR:.1f} 次)'
                    description = f'年化失格频率{AFR:.1f}次/年，{employee_type}阈值{rule["min"]}-{rule["max"]}，需要重点关注。'
                    matched = True
                    break
            else:
                # 只有min的规则（最高阈值）
                if AFR >= rule['min']:
                    coeff = rule['coefficient']
                    tag_level = 'CRITICAL'
                    alert_msg = f'❌ {rule["label"]} (年化 {AFR:.1f} 次)'
                    description = f'当前周期{duration_days}天内失格{fail_count}次，年化等效{AFR:.1f}次/年，超过{employee_type}红线阈值{rule["min"]}次/年。'
                    matched = True
                    break

        if not matched:
            # AFR < 最低阈值
            coeff = 1.0
            tag_level = 'NORMAL'
            alert_msg = '✅ 能力达标'
            description = ''

    # 如果没有失格记录，保持正常
    elif fail_count == 0:
        coeff = 1.0
        tag_level = 'NORMAL'
        alert_msg = '✅ 能力达标'
        description = ''

    # Step 4: 计算最终分数
    final_score = base_score * coeff

    # 映射到前端颜色
    if tag_level == 'CRITICAL':
        status_color = 'RED'
    elif tag_level == 'HIGH_RISK':
        status_color = 'PURPLE'
    elif tag_level == 'WARNING':
        status_color = 'ORANGE'
    elif tag_level == 'NOTICE':
        status_color = 'YELLOW'
    else:
        status_color = 'GREEN'

    return {
        'radar_score': round(final_score, 1),
        'original_score': round(base_score, 1),
        'penalty_coefficient': coeff,
        'stats': {
            'total_ops': total_ops,
            'fail_count': fail_count,
            'duration_days': duration_days
        },
        'risk_alert': {
            'show': fail_count > 0,
            'level': tag_level,
            'text': alert_msg,
            'description': description
        },
        'status_color': status_color,
        'alert_tag': alert_msg
    }


def calculate_learning_ability_monthly(score_curr: float, score_prev: float) -> Dict:
    """
    学习能力评分 - 月度模式 (Algorithm A: Short-Term Sensitivity)

    核心设计：学习能力值 = 现状锚点分 (Position) + 趋势动能分 (Momentum)

    Args:
        score_curr: 本月综合三维得分 (0-100)
        score_prev: 上月综合三维得分 (0-100)，新员工传入 score_curr

    Returns:
        {
            'learning_score': 学习能力分数 (0-100+, 可能超过100),
            'delta': 月度变化量,
            'status_color': 状态颜色 (RED/ORANGE/YELLOW/GREEN/GOLD),
            'alert_tag': 警示标签,
            'tier': 评级 (潜力股/稳健型/懈怠型/高位企稳/低位躺平)
        }
    """
    # Step 1: 计算增量
    delta = score_curr - score_prev

    # Step 2: 计算基础成长分
    # 公式：以本月得分为基准，叠加变化的 1.5 倍权重
    learning_score = score_curr + (delta * 1.5)

    # Step 3: 应用修正逻辑
    tier = '稳健型'
    status_color = 'GREEN'
    alert_tag = '✅ 状态正常'

    # 情形 1：高位企稳 (大师红利)
    if score_curr >= 95 and delta >= -2:
        learning_score = max(100, learning_score)
        tier = '高位企稳'
        status_color = 'GOLD'
        alert_tag = '🏆 顶尖水平 (大师红利)'

    # 情形 2：低位躺平 (差生陷阱)
    elif score_curr < 70 and delta <= 0:
        learning_score = learning_score * 0.8
        tier = '低位躺平'
        status_color = 'RED'
        alert_tag = '❌ 差且无进步 (学习态度有问题)'

    # 情形 3：显著进步
    elif delta > 10:
        tier = '潜力股'
        status_color = 'GOLD'
        alert_tag = f'⭐ 进步神速 (+{delta:.1f}分)'

    # 情形 4：显著退步
    elif delta < -10:
        tier = '懈怠型'
        status_color = 'RED'
        alert_tag = f'⚠️ 严重退步 ({delta:.1f}分)'

    # 情形 5：小幅进步
    elif delta > 0:
        tier = '稳健型'
        status_color = 'GREEN'
        alert_tag = f'✅ 稳中有进 (+{delta:.1f}分)'

    # 情形 6：小幅退步
    elif delta < 0:
        tier = '需关注'
        status_color = 'YELLOW'
        alert_tag = f'⚡ 略有下滑 ({delta:.1f}分)'

    # 限制分数范围（但允许超过100）
    learning_score = max(0, learning_score)

    return {
        'learning_score': round(learning_score, 1),
        'delta': round(delta, 1),
        'slope': 0,  # 月度模式无斜率概念，设为0
        'status_color': status_color,
        'alert_tag': alert_tag,
        'tier': tier
    }


def calculate_learning_ability_longterm(score_list: List[float], config: dict = None, current_three_dim_score: float = None) -> Dict:
    """
    学习能力评分 - 基于线性回归趋势分析

    通过最小二乘法线性回归判断成长趋势，计算学习能力分数

    Args:
        score_list: 过去N个月的三维综合分列表，例如 [85, 86, 88, ..., 92]
                   最少需要2个月数据
        config: 算法配置（可选，默认从数据库读取）
        current_three_dim_score: 当前周期的三维综合分（可选，保留用于向后兼容）

    Returns:
        {
            'learning_score': 学习能力分数 (0-100),
            'slope': 趋势斜率 k (正数表示上升，负数表示下降),
            'average_score': 历史平均分,
            'status_color': 状态颜色,
            'alert_tag': 警示标签,
            'tier': 评级（上升/稳定/下降）
        }
    """
    import numpy as np

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    learning_config = config.get('learning', {
        'potential_threshold': 0.5,
        'decline_threshold': -0.2,
        'decline_penalty': 0.8,
        'slope_amplifier': 10
    })

    # Step 1: 数据验证
    if not score_list or len(score_list) < 2:
        return {
            'learning_score': 0,
            'slope': 0,
            'average_score': 0,
            'status_color': 'GRAY',
            'alert_tag': '⚪ 数据不足',
            'tier': '数据不足'
        }

    # Step 2: 计算线性回归斜率（最小二乘法）
    n = len(score_list)
    x = np.arange(n)
    y = np.array(score_list)

    # 计算斜率 k = (n*Σxy - Σx*Σy) / (n*Σx² - (Σx)²)
    sum_x = np.sum(x)
    sum_y = np.sum(y)
    sum_xy = np.sum(x * y)
    sum_x2 = np.sum(x * x)

    k = (n * sum_xy - sum_x * sum_y) / (n * sum_x2 - sum_x * sum_x) if (n * sum_x2 - sum_x * sum_x) != 0 else 0

    # Step 3: 计算平均分
    average_score = float(np.mean(y))

    # Step 4: 读取配置参数（含 None 检查）
    slope_amplifier = learning_config.get('slope_amplifier', 10)
    if slope_amplifier is None:
        slope_amplifier = 10

    potential_threshold = learning_config.get('potential_threshold', 0.5)
    if potential_threshold is None:
        potential_threshold = 0.5

    decline_threshold = learning_config.get('decline_threshold', -0.2)
    if decline_threshold is None:
        decline_threshold = -0.2

    decline_penalty = learning_config.get('decline_penalty', 0.8)
    if decline_penalty is None:
        decline_penalty = 0.8

    # Step 5: 计算最终得分（简化版：历史平均分 + 趋势加成）
    base_score = average_score
    trend_bonus = k * slope_amplifier
    final_score = base_score + trend_bonus

    # 限制范围
    final_score = max(0, min(100, final_score))

    # Step 6: 根据斜率判断趋势和状态
    if k > potential_threshold:
        tier = '📈 上升趋势'
        status_color = 'GREEN'
        alert_tag = f'表现上升（平均分{average_score:.1f}，斜率{k:.2f}）'
    elif k >= decline_threshold:
        tier = '➡️ 稳定表现'
        status_color = 'BLUE'
        alert_tag = f'表现稳定（平均分{average_score:.1f}，斜率{k:.2f}）'
    else:
        tier = '📉 下降趋势'
        status_color = 'ORANGE'
        alert_tag = f'表现下滑（平均分{average_score:.1f}，斜率{k:.2f}）'

    # Step 7: 返回结果
    return {
        'learning_score': round(final_score, 1),
        'slope': round(k, 3),
        'average_score': round(average_score, 1),
        'status_color': status_color,
        'alert_tag': alert_tag,
        'tier': tier
    }


def calculate_stability_score(
    birth_date: Optional[str],
    work_start_date: Optional[str],
    entry_date: Optional[str],
    certification_date: Optional[str],
    solo_driving_date: Optional[str],
    historical_scores: Optional[Dict[str, List[float]]] = None,
    config: dict = None
) -> Dict:
    """
    职业稳定性综合评分算法（新版）

    评分维度：
    1. 资历维度（60%）：基于年龄、工龄、司龄、取证年限、单独驾驶年限
    2. 表现稳定性维度（40%）：基于过去一年绩效、安全、培训分值的波动度

    Args:
        birth_date: 出生日期 (YYYY-MM-DD)
        work_start_date: 参加工作时间 (YYYY-MM-DD)
        entry_date: 入司时间 (YYYY-MM-DD)
        certification_date: 取证时间 (YYYY-MM-DD)
        solo_driving_date: 单独驾驶时间 (YYYY-MM-DD)
        historical_scores: 过去一年的分数历史，格式：
            {
                'performance': [95.0, 96.0, ...],  # 最多12个月
                'safety': [92.0, 94.0, ...],
                'training': [88.0, 90.0, ...]
            }
        config: 算法配置（可选，默认从数据库读取）

    Returns:
        {
            'stability_score': 最终稳定性分数 (0-100),
            'seniority_score': 资历维度分数 (0-100),
            'volatility_score': 稳定性维度分数 (0-100),
            'metrics': {
                'age_years': 年龄,
                'working_years': 工龄,
                'company_years': 司龄,
                'cert_years': 取证年限,
                'solo_years': 单独驾驶年限,
                'volatility': 综合波动系数
            },
            'status_color': 状态颜色 (RED/ORANGE/GREEN),
            'alert_tag': 警示标签,
            'tier': 评级 (资深稳定/经验丰富/新手期/高波动风险)
        }
    """
    from datetime import datetime
    import numpy as np

    # 读取配置
    if config is None:
        from services.algorithm_config_service import AlgorithmConfigService
        config = AlgorithmConfigService.get_active_config()

    stability_config = config.get('stability', {
        'seniority_weights': {
            'age': 0.15,
            'working_years': 0.20,
            'company_years': 0.25,
            'cert_years': 0.20,
            'solo_years': 0.20
        },
        'seniority_thresholds': {
            'age_cap': 30,  # 年龄满30年算满分
            'working_cap': 20,  # 工龄满20年算满分
            'company_cap': 10,  # 司龄满10年算满分
            'cert_cap': 10,  # 取证满10年算满分
            'solo_cap': 10  # 单独驾驶满10年算满分
        },
        'dimension_weights': {
            'seniority': 0.60,  # 资历维度权重
            'volatility': 0.40   # 稳定性维度权重
        },
        'volatility_penalty': {
            'low_threshold': 5.0,     # 低波动阈值（标准差）
            'high_threshold': 15.0,   # 高波动阈值（标准差）
            'max_penalty': 0.5        # 最大惩罚系数
        }
    })

    now = datetime.now()

    # ==================== 维度1：资历评分（60%） ====================
    seniority_weights = stability_config['seniority_weights']
    seniority_thresholds = stability_config['seniority_thresholds']

    # 1.1 年龄计算
    age_years = 0
    if birth_date:
        try:
            birth = datetime.strptime(birth_date, '%Y-%m-%d')
            age_years = (now - birth).days / 365.25
        except:
            pass
    age_score = min(100, (age_years / seniority_thresholds['age_cap']) * 100)

    # 1.2 工龄计算
    working_years = 0
    if work_start_date:
        try:
            work_start = datetime.strptime(work_start_date, '%Y-%m-%d')
            working_years = (now - work_start).days / 365.25
        except:
            pass
    working_score = min(100, (working_years / seniority_thresholds['working_cap']) * 100)

    # 1.3 司龄计算
    company_years = 0
    if entry_date:
        try:
            entry = datetime.strptime(entry_date, '%Y-%m-%d')
            company_years = (now - entry).days / 365.25
        except:
            pass
    company_score = min(100, (company_years / seniority_thresholds['company_cap']) * 100)

    # 1.4 取证年限计算
    cert_years = 0
    if certification_date:
        try:
            cert = datetime.strptime(certification_date, '%Y-%m-%d')
            cert_years = (now - cert).days / 365.25
        except:
            pass
    cert_score = min(100, (cert_years / seniority_thresholds['cert_cap']) * 100)

    # 1.5 单独驾驶年限计算
    solo_years = 0
    if solo_driving_date:
        try:
            solo = datetime.strptime(solo_driving_date, '%Y-%m-%d')
            solo_years = (now - solo).days / 365.25
        except:
            pass
    solo_score = min(100, (solo_years / seniority_thresholds['solo_cap']) * 100)

    # 计算资历加权分数
    seniority_score = (
        age_score * seniority_weights['age'] +
        working_score * seniority_weights['working_years'] +
        company_score * seniority_weights['company_years'] +
        cert_score * seniority_weights['cert_years'] +
        solo_score * seniority_weights['solo_years']
    )

    # ==================== 维度2：表现稳定性评分（40%） ====================
    volatility_score = 100  # 默认满分（无波动数据时）
    volatility_coefficient = 0

    if historical_scores and any(historical_scores.values()):
        # 计算每个维度的标准差
        std_devs = []

        for dimension in ['performance', 'safety', 'training']:
            scores = historical_scores.get(dimension, [])
            if scores and len(scores) >= 2:
                std_dev = float(np.std(scores))
                std_devs.append(std_dev)

        if std_devs:
            # 综合波动系数：使用平均标准差
            volatility_coefficient = float(np.mean(std_devs))

            # 根据波动系数计算分数
            low_threshold = stability_config['volatility_penalty']['low_threshold']
            high_threshold = stability_config['volatility_penalty']['high_threshold']
            max_penalty = stability_config['volatility_penalty']['max_penalty']

            if volatility_coefficient <= low_threshold:
                # 低波动：满分
                volatility_score = 100
            elif volatility_coefficient >= high_threshold:
                # 高波动：应用最大惩罚
                volatility_score = 100 * (1 - max_penalty)
            else:
                # 中等波动：线性惩罚
                penalty_ratio = (volatility_coefficient - low_threshold) / (high_threshold - low_threshold)
                penalty = max_penalty * penalty_ratio
                volatility_score = 100 * (1 - penalty)

    # ==================== 综合评分 ====================
    dimension_weights = stability_config['dimension_weights']
    final_score = (
        seniority_score * dimension_weights['seniority'] +
        volatility_score * dimension_weights['volatility']
    )

    # ==================== 分级和状态判定 ====================
    # 判定资历等级
    if company_years >= 5 and cert_years >= 5:
        seniority_tier = "资深员工"
    elif company_years >= 2 and cert_years >= 2:
        seniority_tier = "经验员工"
    elif cert_years >= 1:
        seniority_tier = "新手期"
    else:
        seniority_tier = "新员工"

    # 判定稳定性等级
    if volatility_coefficient == 0:
        volatility_tier = "无历史数据"
    elif volatility_coefficient <= low_threshold:
        volatility_tier = "表现稳定"
    elif volatility_coefficient <= high_threshold:
        volatility_tier = "波动适中"
    else:
        volatility_tier = "高波动风险"

    # 综合评级
    if final_score >= 85:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'GREEN'
        alert_tag = '✅ 稳定可靠'
    elif final_score >= 70:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'GREEN'
        alert_tag = '✅ 基本稳定'
    elif final_score >= 50:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'ORANGE'
        alert_tag = '⚠️ 稳定性一般'
    else:
        tier = f"{seniority_tier}·{volatility_tier}"
        status_color = 'RED'
        alert_tag = '⛔ 不稳定'

    return {
        'stability_score': round(final_score, 1),
        'seniority_score': round(seniority_score, 1),
        'volatility_score': round(volatility_score, 1),
        'metrics': {
            'age_years': round(age_years, 1),
            'working_years': round(working_years, 1),
            'company_years': round(company_years, 1),
            'cert_years': round(cert_years, 1),
            'solo_years': round(solo_years, 1),
            'volatility': round(volatility_coefficient, 2)
        },
        'status_color': status_color,
        'alert_tag': alert_tag,
        'tier': tier
    }


def _parse_date_string(value: Optional[str]) -> Optional[date]:
    """解析日期字符串为date对象"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    fmts = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y%m%d",
        "%Y-%m",
        "%Y/%m",
        "%Y.%m",
        "%Y%m",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(raw, fmt)
            if fmt in {"%Y-%m", "%Y/%m", "%Y.%m", "%Y%m"}:
                dt = dt.replace(day=1)
            return dt.date()
        except ValueError:
            continue
    return None


def _normalize_date_to_str(value: Optional[str]) -> Optional[str]:
    """标准化日期为字符串"""
    parsed = _parse_date_string(value)
    return parsed.strftime("%Y-%m-%d") if parsed else None


def _calculate_age(birth_date: Optional[str]) -> Optional[int]:
    """计算年龄"""
    parsed = _parse_date_string(birth_date)
    if not parsed:
        return None
    today = date.today()
    years = today.year - parsed.year
    if (today.month, today.day) < (parsed.month, parsed.day):
        years -= 1
    return max(years, 0)


def _calculate_years_since(date_str: Optional[str]) -> Optional[float]:
    """计算从指定日期到今天的年数"""
    parsed = _parse_date_string(date_str)
    if not parsed:
        return None
    today = date.today()
    if parsed > today:
        return 0.0
    years = (today - parsed).days / 365.25
    return round(years, 1)


def _serialize_person(row: sqlite3.Row) -> Dict:
    """序列化人员数据，添加计算字段"""
    data = dict(row)
    data["age"] = _calculate_age(data.get("birth_date"))
    data["working_years"] = _calculate_years_since(data.get("work_start_date"))
    data["tenure_years"] = _calculate_years_since(data.get("entry_date"))
    return data


def _build_personnel_charts(rows: List[Dict]) -> Dict:
    """构建人员统计图表数据"""
    # 年龄分布
    age_labels = ["25岁及以下", "26-35岁", "36-45岁", "46岁及以上"]
    age_counts = [0, 0, 0, 0]
    for row in rows:
        age = row.get("age")
        if age is None:
            continue
        if age <= 25:
            age_counts[0] += 1
        elif 26 <= age <= 35:
            age_counts[1] += 1
        elif 36 <= age <= 45:
            age_counts[2] += 1
        else:
            age_counts[3] += 1

    # 学历分布
    education_counter = Counter(
        row.get("education") or "未填写" for row in rows
    )
    education_labels = list(education_counter.keys())
    education_counts = [education_counter[label] for label in education_labels]

    # 工龄分布
    tenure_labels = ["1年以下", "1-3年", "3-5年", "5-10年", "10年以上"]
    tenure_counts = [0, 0, 0, 0, 0]
    for row in rows:
        tenure = row.get("tenure_years")
        if tenure is None:
            continue
        if tenure < 1:
            tenure_counts[0] += 1
        elif 1 <= tenure < 3:
            tenure_counts[1] += 1
        elif 3 <= tenure < 5:
            tenure_counts[2] += 1
        elif 5 <= tenure < 10:
            tenure_counts[3] += 1
        else:
            tenure_counts[4] += 1

    return {
        "age": {"labels": age_labels, "values": age_counts},
        "education": {"labels": education_labels, "values": education_counts},
        "tenure": {"labels": tenure_labels, "values": tenure_counts},
    }


# ==================== 数据库访问函数 ====================

def list_personnel():
    """列出所有可访问的人员"""
    from flask import session
    user_role = session.get('role', 'user')

    conn = get_db()
    cur = conn.cursor()

    # 管理员可以看到所有员工，其他角色只能看到可访问部门的员工
    if user_role == 'admin':
        query = """
            SELECT e.emp_no, e.name, e.department_id, d.name as department_name,
                   e.class_name, e.position, e.birth_date, e.certification_date,
                   e.solo_driving_date, e.marital_status, e.hometown,
                   e.political_status, e.education, e.graduation_school,
                   e.work_start_date, e.entry_date, e.specialty
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            ORDER BY CAST(e.emp_no as INTEGER)
        """
        try:
            cur.execute(query)
        except sqlite3.OperationalError:
            cur.execute(query.replace("CAST(e.emp_no as INTEGER)", "e.emp_no"))
    else:
        accessible_dept_ids = get_accessible_department_ids()
        if not accessible_dept_ids:
            return []

        placeholders = ','.join('?' * len(accessible_dept_ids))
        query = f"""
            SELECT e.emp_no, e.name, e.department_id, d.name as department_name,
                   e.class_name, e.position, e.birth_date, e.certification_date,
                   e.solo_driving_date, e.marital_status, e.hometown,
                   e.political_status, e.education, e.graduation_school,
                   e.work_start_date, e.entry_date, e.specialty
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            WHERE e.department_id IN ({placeholders})
            ORDER BY CAST(e.emp_no as INTEGER)
        """
        try:
            cur.execute(query, accessible_dept_ids)
        except sqlite3.OperationalError:
            cur.execute(
                query.replace("CAST(e.emp_no as INTEGER)", "e.emp_no"),
                accessible_dept_ids,
            )

    rows = cur.fetchall()
    result = []
    for row in rows:
        person_dict = _serialize_person(row)
        # 添加计算字段
        if person_dict.get('certification_date'):
            person_dict['certification_years'] = calculate_years_from_date(person_dict['certification_date'])
        else:
            person_dict['certification_years'] = None

        if person_dict.get('solo_driving_date'):
            person_dict['solo_driving_years'] = calculate_years_from_date(person_dict['solo_driving_date'])
        else:
            person_dict['solo_driving_years'] = None

        result.append(person_dict)

    return result


def get_personnel(emp_no: str) -> Optional[Dict]:
    """获取指定工号的人员信息"""
    uid = require_user_id()

    # 🔒 权限检查: 非管理员需要验证是否有权访问该员工
    from flask import session
    user_role = session.get('role', 'user')
    if user_role != 'admin':
        if not validate_employee_access(emp_no):
            return None

    conn = get_db()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(
        """
        SELECT e.emp_no, e.name, e.department_id, d.name as department_name,
               e.class_name, e.position, e.birth_date, e.certification_date,
               e.solo_driving_date, e.marital_status, e.hometown,
               e.political_status, e.education, e.graduation_school,
               e.work_start_date, e.entry_date, e.specialty, e.created_at
        FROM employees e
        LEFT JOIN departments d ON e.department_id = d.id
        WHERE e.emp_no=?
        """,
        (emp_no,),
    )
    row = cur.fetchone()
    if not row:
        return None

    person_dict = _serialize_person(row)
    # 添加计算字段
    if person_dict.get('certification_date'):
        person_dict['certification_years'] = calculate_years_from_date(person_dict['certification_date'])
    if person_dict.get('solo_driving_date'):
        person_dict['solo_driving_years'] = calculate_years_from_date(person_dict['solo_driving_date'])

    return person_dict


def _sanitize_person_payload(data: Dict[str, Optional[str]]) -> Dict[str, Optional[str]]:
    """清理和标准化人员数据"""
    sanitized: Dict[str, Optional[str]] = {}
    for field in PERSONNEL_DB_COLUMNS + ["emp_no", "name"]:
        if field == "emp_no":
            value = str(data.get(field) or "").strip()
            sanitized[field] = value or None
            continue
        raw_val = data.get(field)
        if raw_val is None:
            sanitized[field] = None
            continue
        if field in PERSONNEL_DATE_FIELDS:
            sanitized[field] = _normalize_date_to_str(raw_val)
        else:
            sanitized[field] = str(raw_val).strip() or None
    return sanitized


def upsert_personnel(data: Dict[str, Optional[str]]) -> bool:
    """插入或更新人员信息"""
    payload = _sanitize_person_payload(data)
    emp_no = payload.get("emp_no")
    name = payload.get("name")
    department_id = payload.get("department_id")

    if not emp_no or not name:
        return False

    # department_id是必填项，如果没有提供则返回False
    if department_id is None or department_id == "":
        return False

    # 转换department_id为整数
    try:
        department_id = int(department_id)
    except (ValueError, TypeError):
        return False

    uid = require_user_id()
    conn = get_db()
    cur = conn.cursor()

    # 注意: UNIQUE约束是emp_no（全局唯一），数据以department_id为基准隔离
    columns = ["emp_no", "name", "created_by", "department_id"] + [col for col in PERSONNEL_DB_COLUMNS if col != "department_id"]
    values = [emp_no, name, uid, department_id] + [payload.get(col) for col in PERSONNEL_DB_COLUMNS if col != "department_id"]
    update_clause = ", ".join(
        f"{col}=excluded.{col}" for col in ["name", "department_id"] + [col for col in PERSONNEL_DB_COLUMNS if col != "department_id"]
    )
    cur.execute(
        f"""
        INSERT INTO employees ({", ".join(columns)})
        VALUES ({", ".join("?" for _ in columns)})
        ON CONFLICT(emp_no) DO UPDATE SET {update_clause}
        """,
        values,
    )
    conn.commit()
    return True


def bulk_import_personnel(records: List[Dict[str, Optional[str]]]) -> int:
    """批量导入人员信息"""
    imported = 0
    for record in records:
        if upsert_personnel(record):
            imported += 1
    return imported


def update_personnel_field(emp_no: str, field: str, value: Optional[str]) -> bool:
    """更新人员的单个字段"""
    if field not in {"name", *PERSONNEL_DB_COLUMNS}:
        return False

    # 🔒 权限检查: 非管理员需要验证是否有权修改该员工
    from flask import session
    user_role = session.get('role', 'user')
    if user_role != 'admin':
        if not validate_employee_access(emp_no):
            return False

    payload = _sanitize_person_payload({field: value})
    uid = require_user_id()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        f"""
        UPDATE employees
        SET {field} = ?
        WHERE emp_no=?
        """,
        (payload.get(field), emp_no),
    )
    conn.commit()
    affected = cur.rowcount > 0
    return affected


def delete_employee(emp_no):
    """删除员工"""
    uid = require_user_id()

    # 🔒 权限检查: 非管理员需要验证是否有权删除该员工
    from flask import session
    user_role = session.get('role', 'user')
    if user_role != 'admin':
        if not validate_employee_access(emp_no):
            return False

    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM employees WHERE emp_no=?", (emp_no,))
    conn.commit()
    return True


# ==================== 路由处理 ====================

@personnel_bp.route('/', methods=['GET', 'POST'])
@login_required
def index():
    """人员管理首页"""
    if request.method == 'POST':
        # 🔒 权限检查: 创建/更新员工需要管理员权限
        from flask import session
        user_role = session.get('role', 'user')
        if user_role not in ['admin', 'manager']:
            flash("您没有权限执行此操作，需要部门管理员或系统管理员权限", "danger")
            return redirect(url_for("personnel.index"))

        form_payload = {}
        for field in PERSONNEL_FIELD_SCHEME:
            key = field["name"]
            if field["input_type"] == "textarea":
                form_payload[key] = request.form.get(key, "")
            else:
                form_payload[key] = request.form.get(key, "").strip()
        saved = upsert_personnel(form_payload)
        if saved:
            flash("人员信息已保存。", "success")
        else:
            flash("请填写有效的工号和姓名。", "warning")
        return redirect(url_for("personnel.index"))

    rows = list_personnel()
    accessible_departments = get_accessible_departments()

    return render_template(
        "personnel.html",
        title=f"人员管理 | {APP_TITLE}",
        rows=rows,
        field_scheme=PERSONNEL_FIELD_SCHEME,
        select_options=PERSONNEL_SELECT_OPTIONS,
        accessible_departments=accessible_departments,
    )


@personnel_bp.route('/template')
@login_required
def template():
    """下载人员导入模板"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人员导入模板"

    headers = [field["label"] for field in PERSONNEL_FIELD_SCHEME]
    sheet.append(headers)

    examples = {
        "emp_no": "1001",
        "name": "张三",
        "class_name": "一班",
        "position": "班长",
        "birth_date": "1990-01-01",
        "marital_status": "已婚",
        "hometown": "江苏南京",
        "political_status": "群众",
        "education": "本科",
        "graduation_school": "某某大学",
        "work_start_date": "2012-07-01",
        "entry_date": "2018-03-15",
        "specialty": "摄影、篮球",
    }
    sheet.append([examples.get(field["name"], "") for field in PERSONNEL_FIELD_SCHEME])

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"personnel_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@personnel_bp.route('/import', methods=['POST'])
@manager_required
def import_data():
    """批量导入人员数据"""
    file_obj = request.files.get("file")
    if not file_obj or file_obj.filename == "":
        flash("请选择包含花名册数据的 Excel 文件。", "warning")
        return redirect(url_for("personnel.index"))
    ext = file_obj.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"xlsx"}:
        flash("目前仅支持上传 .xlsx 文件。", "warning")
        return redirect(url_for("personnel.index"))
    try:
        workbook = load_workbook(file_obj, data_only=True)
        sheet = workbook.active
    except Exception as exc:  # noqa: BLE001
        flash(f"无法读取 Excel 文件：{exc}", "danger")
        return redirect(url_for("personnel.index"))

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        flash("Excel 文件为空。", "warning")
        return redirect(url_for("personnel.index"))

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    field_map = [PERSONNEL_IMPORT_HEADER_MAP.get(header) for header in headers]

    if "emp_no" not in field_map or "name" not in field_map:
        flash('Excel 首行必须包含"工号"与"姓名"列。', "warning")
        return redirect(url_for("personnel.index"))

    # 获取部门映射，用于处理Excel中的部门信息
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM departments")
    dept_name_map = {row['name']: row['id'] for row in cur.fetchall()}

    # 获取当前用户可访问的部门ID列表（用于权限验证）
    accessible_dept_ids = get_accessible_department_ids()

    records: List[Dict[str, Optional[str]]] = []
    skipped_no_dept = 0
    skipped_no_permission = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: Dict[str, Optional[str]] = {}
        for idx, cell in enumerate(row):
            field = field_map[idx] if idx < len(field_map) else None
            if not field:
                continue
            record[field] = cell
            
        # 处理部门ID：支持名称匹配
        raw_dept = record.get('department_id')
        final_dept_id = None

        if raw_dept:
            raw_dept_str = str(raw_dept).strip()
            if raw_dept_str.isdigit():
                final_dept_id = int(raw_dept_str)
            elif raw_dept_str in dept_name_map:
                final_dept_id = dept_name_map[raw_dept_str]

        if not final_dept_id:
            # 未填写部门或部门无效
            skipped_no_dept += 1
        elif final_dept_id not in accessible_dept_ids:
            # 部门存在但无权限导入到该部门
            skipped_no_permission += 1
        else:
            # 部门有效且有权限
            record['department_id'] = str(final_dept_id)
            records.append(record)

    if not records:
        msg_parts = ["未导入任何数据。"]
        if skipped_no_dept > 0:
            msg_parts.append(f"{skipped_no_dept} 条记录因未填写部门或部门无效被跳过。")
        if skipped_no_permission > 0:
            msg_parts.append(f"{skipped_no_permission} 条记录因无权限导入到该部门被跳过。")
        if not skipped_no_dept and not skipped_no_permission:
            msg_parts.append("未识别到任何有效行。")
        flash(" ".join(msg_parts), "warning")

        # 记录失败的导入操作
        log_import_operation(
            module='personnel',
            operation='import',
            file_name=file_obj.filename,
            total_rows=skipped_no_dept + skipped_no_permission,
            success_rows=0,
            failed_rows=0,
            skipped_rows=skipped_no_dept + skipped_no_permission,
            error_message=" ".join(msg_parts),
            import_details={
                'skipped_no_dept': skipped_no_dept,
                'skipped_no_permission': skipped_no_permission
            }
        )
        return redirect(url_for("personnel.index"))

    imported = bulk_import_personnel(records)

    # 计算总行数
    total_rows = len(records) + skipped_no_dept + skipped_no_permission

    # 构建提示消息
    msg = f"已导入/更新 {imported} 名员工信息。"
    msg_parts = []
    if skipped_no_dept > 0:
        msg_parts.append(f"{skipped_no_dept} 条记录因未填写部门或部门无效被跳过")
    if skipped_no_permission > 0:
        msg_parts.append(f"{skipped_no_permission} 条记录因无权限导入到该部门被跳过")

    if msg_parts:
        msg += " 另有 " + "、".join(msg_parts) + "。"
        flash_type = "warning"
    else:
        flash_type = "success"

    flash(msg, flash_type)

    # 记录导入操作日志
    log_import_operation(
        module='personnel',
        operation='import',
        file_name=file_obj.filename,
        total_rows=total_rows,
        success_rows=imported,
        failed_rows=0,
        skipped_rows=skipped_no_dept + skipped_no_permission,
        import_details={
            'imported': imported,
            'skipped_no_dept': skipped_no_dept,
            'skipped_no_permission': skipped_no_permission,
            'accessible_departments': len(accessible_dept_ids)
        }
    )

    return redirect(url_for("personnel.index"))


@personnel_bp.route('/<emp_no>')
@login_required
def preview(emp_no):
    """查看人员详情"""
    person = get_personnel(emp_no)
    if not person:
        flash("未找到该员工。", "warning")
        return redirect(url_for("personnel.index"))
    return render_template(
        "personnel_preview.html",
        title=f"{person.get('name', '')} | 人员档案 · {APP_TITLE}",
        person=person,
        field_scheme=PERSONNEL_FIELD_SCHEME,
        select_options=PERSONNEL_SELECT_OPTIONS,
    )


@personnel_bp.route('/<emp_no>/update', methods=['POST'])
@login_required
@manager_required
def update(emp_no):
    """更新人员信息字段（仅限部门管理员及以上权限）"""
    payload = request.get_json(silent=True) or request.form
    field = (payload.get("field") or "").strip()
    value = payload.get("value")
    if field in PERSONNEL_DATE_FIELDS and isinstance(value, str):
        value = value.strip()
    if not field:
        return jsonify({"ok": False, "message": "未指定字段"}), 400
    updated = update_personnel_field(emp_no, field, value)
    if not updated:
        return jsonify({"ok": False, "message": "更新失败或字段不受支持"}), 400
    person = get_personnel(emp_no)
    return jsonify({"ok": True, "person": person})


@personnel_bp.route('/batch-delete', methods=['POST'])
@login_required
@manager_required
def batch_delete():
    """批量删除员工（仅限部门管理员及以上权限）"""
    emp_nos = request.form.getlist('emp_nos')

    if not emp_nos:
        flash("未选择要删除的员工", "warning")
        return redirect(url_for("personnel.index"))

    uid = require_user_id()
    from flask import session
    user_role = session.get('role', 'user')

    conn = get_db()
    cur = conn.cursor()

    deleted_count = 0
    skipped_count = 0
    for emp_no in emp_nos:
        emp_no = emp_no.strip()
        if emp_no:
            # 🔒 权限检查: 非管理员需要验证是否有权删除每个员工
            if user_role != 'admin':
                if not validate_employee_access(emp_no):
                    skipped_count += 1
                    continue

            cur.execute("DELETE FROM employees WHERE emp_no=?", (emp_no,))
            if cur.rowcount > 0:
                deleted_count += 1

    conn.commit()

    if deleted_count > 0:
        message = f"成功删除 {deleted_count} 名员工"
        if skipped_count > 0:
            message += f"，跳过 {skipped_count} 名无权删除的员工"
        flash(message, "success")
    elif skipped_count > 0:
        flash(f"跳过 {skipped_count} 名无权删除的员工", "warning")
    else:
        flash("未删除任何员工", "info")

    return redirect(url_for("personnel.index"))


@personnel_bp.route('/employees')
@login_required
def employees_legacy_redirect():
    """旧版employees路由重定向"""
    flash("花名册入口已升级为人员管理，请使用新页面。", "info")
    return redirect(url_for("personnel.index"))


@personnel_bp.route('/dashboard')
@login_required
def dashboard():
    """人员工作台首页"""
    feature_cards = [
        {
            "title": "人员管理",
            "description": "查看和管理员工档案信息，支持批量导入导出",
            "endpoint": "personnel.index"
        },
        {
            "title": "数据分析",
            "description": "可视化分析人员结构、班组战力、经验分布等关键指标",
            "endpoint": "personnel.analytics"
        },
        {
            "title": "能力画像",
            "description": "整合人员、培训、安全、绩效数据的五维能力评估（按权限分级）",
            "endpoint": "personnel.capability_profile"
        },
    ]
    return render_template(
        "personnel_dashboard.html",
        title=f"人员工作台 | {APP_TITLE}",
        feature_cards=feature_cards
    )


@personnel_bp.route('/analytics')
@login_required
def analytics():
    """人员数据分析页面"""
    return render_template(
        "personnel_analytics.html",
        title=f"人员数据分析 | {APP_TITLE}"
    )


@personnel_bp.route('/api/analytics-data')
@login_required
def api_analytics_data():
    """获取人员分析数据API"""
    rows = list_personnel()

    # 岗位筛选：只统计电客车司机，排除副队长和队长
    def is_driver(row):
        position = (row.get("position") or "").strip()
        # 排除副队长和队长
        if "队长" in position:
            return False
        # 只要包含"司机"就算
        return "司机" in position

    # 除了政治面貌统计，其他都只统计司机
    driver_rows = [row for row in rows if is_driver(row)]

    # 1. 安全风险等级分布 - 按入司后单独驾驶年限分级
    risk_levels = {"新手(<1年)": 0, "成长(1-3年)": 0, "熟练(3-5年)": 0, "资深(≥5年)": 0, "未知": 0}
    for row in driver_rows:
        solo_years = row.get("solo_driving_years")
        if solo_years is None:
            risk_levels["未知"] += 1
        elif solo_years < 1:
            risk_levels["新手(<1年)"] += 1
        elif 1 <= solo_years < 3:
            risk_levels["成长(1-3年)"] += 1
        elif 3 <= solo_years < 5:
            risk_levels["熟练(3-5年)"] += 1
        else:
            risk_levels["资深(≥5年)"] += 1

    # 2. 部门战力雷达图 - 各部门的平均司龄、驾龄、取证年限（只统计司机）
    # 获取当前用户可访问的部门列表
    accessible_depts = get_accessible_departments()

    # 获取所有底层部门（没有子部门的部门）
    conn = get_db()
    cur = conn.cursor()

    # 找出所有可访问部门中的底层部门
    accessible_dept_ids = [dept['id'] for dept in accessible_depts]
    if not accessible_dept_ids:
        team_power = []
    else:
        # 查询每个部门是否有子部门
        placeholders = ','.join('?' * len(accessible_dept_ids))
        cur.execute(f"""
            SELECT d.id, d.name, d.level,
                   CASE
                       WHEN EXISTS(SELECT 1 FROM departments child WHERE child.parent_id = d.id)
                       THEN 1 ELSE 0
                   END as has_children
            FROM departments d
            WHERE d.id IN ({placeholders})
            ORDER BY d.level, d.name
        """, accessible_dept_ids)

        dept_info = {row['id']: dict(row) for row in cur.fetchall()}

        # 对于最底层用户，只显示自己部门；对于上级用户，显示所有下级底层部门
        user_dept_info = get_user_department()
        if user_dept_info and user_dept_info['department_id']:
            user_dept_id = user_dept_info['department_id']
            # 检查用户部门是否是底层部门
            if user_dept_id in dept_info and dept_info[user_dept_id]['has_children'] == 0:
                # 用户是底层部门，只显示自己部门
                display_dept_ids = [user_dept_id]
            else:
                # 用户是上级部门，显示所有可访问的底层部门
                display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]
        else:
            # 管理员或无部门用户，显示所有底层部门
            display_dept_ids = [dept_id for dept_id, info in dept_info.items() if info['has_children'] == 0]

        # 按部门统计司机数据
        dept_stats = {}
        for row in driver_rows:
            dept_id = row.get("department_id")
            if dept_id not in display_dept_ids:
                continue

            if dept_id not in dept_stats:
                dept_stats[dept_id] = {
                    "name": dept_info.get(dept_id, {}).get('name', '未知部门'),
                    "tenure_years": [],
                    "solo_driving_years": [],
                    "certification_years": []
                }

            if row.get("tenure_years") is not None:
                dept_stats[dept_id]["tenure_years"].append(row["tenure_years"])
            if row.get("solo_driving_years") is not None:
                dept_stats[dept_id]["solo_driving_years"].append(row["solo_driving_years"])
            if row.get("certification_years") is not None:
                dept_stats[dept_id]["certification_years"].append(row["certification_years"])

        team_power = []
        for dept_id, stats in dept_stats.items():
            avg_tenure = sum(stats["tenure_years"]) / len(stats["tenure_years"]) if stats["tenure_years"] else 0
            avg_solo = sum(stats["solo_driving_years"]) / len(stats["solo_driving_years"]) if stats["solo_driving_years"] else 0
            avg_cert = sum(stats["certification_years"]) / len(stats["certification_years"]) if stats["certification_years"] else 0

            team_power.append({
                "team": stats["name"],
                "avg_tenure": round(avg_tenure, 1),
                "avg_solo": round(avg_solo, 1),
                "avg_cert": round(avg_cert, 1),
                "member_count": len([r for r in driver_rows if r.get("department_id") == dept_id])
            })

    # 3. 经验溢出分析 - 散点图数据（只统计司机）
    experience_scatter = []
    for row in driver_rows:
        cert_years = row.get("certification_years")
        solo_years = row.get("solo_driving_years")
        if cert_years is not None and solo_years is not None:
            experience_scatter.append({
                "name": row.get("name"),
                "emp_no": row.get("emp_no"),
                "cert_years": round(cert_years, 1),
                "solo_years": round(solo_years, 1),
                # 分类：准师傅(取证久但单驾短)、资深师傅(两项都高)、新手
                "category": _categorize_experience(cert_years, solo_years)
            })

    # 4. 排班压力预警 - 籍贯分布（只统计司机）+ 政治面貌统计（统计所有人）
    hometown_stats = {}
    political_stats = {"中共党员": 0, "中共预备党员": 0, "共青团员": 0, "群众": 0, "其它": 0}

    # 籍贯统计只统计司机
    for row in driver_rows:
        hometown = row.get("hometown") or "未填写"
        # 河南省内细分到市/县，省外只显示省份
        location = _extract_location(hometown)
        hometown_stats[location] = hometown_stats.get(location, 0) + 1

    # 政治面貌统计所有人员
    for row in rows:
        political = row.get("political_status") or "未填写"
        if political in political_stats:
            political_stats[political] += 1
        else:
            political_stats["其它"] += 1

    # 5. 职业稳定性分析 - 司龄 vs 工龄散点图（只统计司机）
    stability_scatter = []
    for row in driver_rows:
        tenure = row.get("tenure_years")
        working = row.get("working_years")
        if tenure is not None and working is not None:
            stability_scatter.append({
                "name": row.get("name"),
                "emp_no": row.get("emp_no"),
                "tenure": round(tenure, 1),
                "working": round(working, 1),
                # 分类：应届入职、社招新员工、社招老员工
                "category": _categorize_stability(tenure, working)
            })

    return jsonify({
        "risk_distribution": risk_levels,
        "team_power": team_power,
        "experience_scatter": experience_scatter,
        "hometown_stats": hometown_stats,
        "political_stats": political_stats,
        "stability_scatter": stability_scatter,
        "total_count": len(rows),
        "driver_count": len(driver_rows)
    })


def _categorize_experience(cert_years: float, solo_years: float) -> str:
    """分类经验等级"""
    if cert_years >= 5 and solo_years < 3:
        return "准师傅"  # 取证很久但单驾时间较短
    elif cert_years >= 5 and solo_years >= 5:
        return "资深师傅"  # 两项指标都高
    elif cert_years < 2:
        return "新手"
    else:
        return "普通"


def _categorize_stability(tenure: float, working: float) -> str:
    """分类职业稳定性

    Args:
        tenure: 司龄（在本单位工作年限）
        working: 工龄（总工作年限）

    Returns:
        分类标签：应届入职、社招(新)、社招(老)
    """
    work_exp_diff = working - tenure  # 入职前的工作经验

    if work_exp_diff < 1:
        # 工龄和司龄相近，基本是应届生或毕业后很快入职
        return "应届入职"
    elif tenure < 3:
        # 有工作经验，但在本单位时间不长
        return "社招(新)"
    else:
        # 有工作经验，且在本单位时间较长
        return "社招(老)"


def _extract_location(hometown: str) -> str:
    """提取地域信息

    河南省内细分到市/县，其他省份只显示省外或省份名称

    Args:
        hometown: 籍贯字符串，如"河南郑州"、"河南省洛阳市"、"江苏南京"等

    Returns:
        地域标签：河南省内返回市/县名，省外返回省份名或"省外"
    """
    if not hometown or hometown == "未填写":
        return "未填写"

    hometown = hometown.strip()

    # 河南省内的地级市和县
    henan_cities = [
        "郑州", "开封", "洛阳", "平顶山", "安阳", "鹤壁",
        "新乡", "焦作", "濮阳", "许昌", "漯河", "三门峡",
        "南阳", "商丘", "信阳", "周口", "驻马店", "济源"
    ]

    # 常见县级市/县（可根据实际情况扩展）
    henan_counties = [
        "巩义", "荥阳", "新密", "新郑", "登封", "中牟",
        "兰考", "杞县", "通许", "尉氏", "偃师", "孟津",
        "新安", "栾川", "嵩县", "汝阳", "宜阳", "洛宁",
        "伊川", "汝州", "舞钢", "林州", "卫辉", "辉县",
        "沁阳", "孟州", "禹州", "长葛", "义马", "灵宝",
        "永城", "项城", "邓州", "固始", "鹿邑", "新蔡"
    ]

    # 检查是否为河南省内
    is_henan = False
    if "河南" in hometown:
        is_henan = True
    else:
        # 如果没有明确写"河南"，但包含河南的市/县名，也认为是河南
        for city in henan_cities + henan_counties:
            if city in hometown:
                is_henan = True
                break

    if is_henan:
        # 河南省内，提取市/县名
        # 优先匹配县级市/县（更具体）
        for county in henan_counties:
            if county in hometown:
                return f"河南·{county}"

        # 再匹配地级市
        for city in henan_cities:
            if city in hometown:
                return f"河南·{city}"

        # 如果只写了"河南"，返回"河南·未详"
        return "河南·未详"

    else:
        # 非河南省，提取省份
        provinces = [
            "北京", "天津", "上海", "重庆",
            "河北", "山西", "辽宁", "吉林", "黑龙江",
            "江苏", "浙江", "安徽", "福建", "江西", "山东",
            "湖北", "湖南", "广东", "海南",
            "四川", "贵州", "云南", "陕西", "甘肃",
            "青海", "台湾", "内蒙古", "广西", "西藏",
            "宁夏", "新疆", "香港", "澳门"
        ]

        for province in provinces:
            if hometown.startswith(province) or province in hometown:
                return f"省外·{province}"

        # 如果无法识别，返回"省外·其他"
        return "省外·其他"


# ==================== 个人综合能力画像 API ====================

@personnel_bp.route('/capability-profile')
@login_required
def capability_profile():
    """个人综合能力画像页面"""
    return render_template('personnel_capability_profile.html', title='个人综合能力画像')


@personnel_bp.route('/api/key-personnel-config')
@login_required
def api_key_personnel_config():
    """API: 获取关键人员配置参数（供前端动态显示使用）"""
    from services.algorithm_config_service import AlgorithmConfigService

    try:
        algo_config = AlgorithmConfigService.get_active_config()
        key_personnel_config = algo_config.get('key_personnel', {})

        return jsonify({
            'success': True,
            'config': {
                'comprehensive_threshold': key_personnel_config.get('comprehensive_threshold', 75),
                'monthly_violation_threshold': key_personnel_config.get('monthly_violation_threshold', 3)
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'config': {
                'comprehensive_threshold': 75,
                'monthly_violation_threshold': 3
            }
        })


@personnel_bp.route('/api/students-list')
@login_required
def api_students_list():
    """API: 获取人员列表及综合评分（带权限过滤和关键人员标记）"""
    from datetime import datetime
    from blueprints.safety import extract_score_from_assessment

    conn = get_db()
    cur = conn.cursor()

    # 获取筛选参数
    start_date = request.args.get('start_date')  # 格式：YYYY-MM
    end_date = request.args.get('end_date')      # 格式：YYYY-MM
    department_filter = request.args.get('department')
    position_filter = request.args.get('position')

    # 如果没有指定日期筛选，默认使用当月（1号到今天）
    if not start_date and not end_date:
        current_month = datetime.now().strftime('%Y-%m')
        start_date = current_month
        end_date = current_month

    # 获取当前月份（用于关键人员标记）
    current_month = datetime.now().strftime('%Y-%m')

    # 读取算法配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    score_weights = algo_config['comprehensive']['score_weights']
    key_personnel_config = algo_config['key_personnel']

    # 兼容 sqlite3.Row 和 dict 两种类型的辅助函数
    def safe_get(obj, key, default=None):
        if isinstance(obj, dict):
            return obj.get(key, default)
        else:
            try:
                return obj[key] if obj[key] is not None else default
            except (KeyError, IndexError):
                return default

    # 使用现有的 list_personnel() 函数获取权限过滤后的人员列表
    rows = list_personnel()

    # 应用部门和岗位筛选
    if department_filter:
        rows = [r for r in rows if safe_get(r, 'department_name') == department_filter]

    if position_filter:
        position_filter_lower = position_filter.lower()
        rows = [r for r in rows if position_filter_lower in (safe_get(r, 'position') or '').lower()]

    students = []
    for row in rows:
        emp_no = safe_get(row, 'emp_no')
        emp_name = safe_get(row, 'name')
        dept_id = safe_get(row, 'department_id')

        # 计算取证年限（用于培训和稳定性算法）
        cert_date = safe_get(row, 'certification_date')
        cert_years = calculate_years_from_date(cert_date) if cert_date else None

        # 获取部门名称
        if dept_id:
            cur.execute("SELECT name FROM departments WHERE id = ?", (dept_id,))
            dept_row = cur.fetchone()
            dept_name = dept_row[0] if dept_row else None
        else:
            dept_name = None

        # 1. 培训能力（使用高级评分算法，应用日期筛选）
        training_query = """
            SELECT score, is_qualified, is_disqualified, training_date
            FROM training_records
            WHERE emp_no = ?
        """
        training_params = [emp_no]

        if start_date:
            training_query += " AND strftime('%Y-%m', training_date) >= ?"
            training_params.append(start_date)

        if end_date:
            training_query += " AND strftime('%Y-%m', training_date) <= ?"
            training_params.append(end_date)

        training_query += " ORDER BY training_date ASC"
        cur.execute(training_query, training_params)
        training_records_list = cur.fetchall()

        # 计算统计周期天数
        if start_date and end_date and start_date == end_date:
            duration_days = 30
        elif start_date and end_date:
            try:
                start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
                import calendar
                end_year, end_month = int(end_date.split('-')[0]), int(end_date.split('-')[1])
                last_day = calendar.monthrange(end_year, end_month)[1]
                end_dt = end_dt.replace(day=last_day)
                duration_days = max(1, (end_dt - start_dt).days + 1)
            except:
                duration_days = 30
        else:
            duration_days = 30

        # 使用新的评分算法
        training_result = calculate_training_score_with_penalty(training_records_list, duration_days, cert_years, algo_config)
        training_score = training_result['radar_score']

        # 2. 安全意识（使用双轨评分模型）
        # 构建日期筛选条件
        safety_query = """
            SELECT assessment, inspection_date
            FROM safety_inspection_records
            WHERE inspected_person = ?
        """
        safety_params = [emp_name]

        if start_date:
            safety_query += " AND strftime('%Y-%m', inspection_date) >= ?"
            safety_params.append(start_date)

        if end_date:
            safety_query += " AND strftime('%Y-%m', inspection_date) <= ?"
            safety_params.append(end_date)

        safety_query += " ORDER BY inspection_date ASC"
        cur.execute(safety_query, safety_params)
        safety_rows = cur.fetchall()

        # 收集所有违规扣分
        violations_list = []
        for s_row in safety_rows:
            assessment = s_row[0]
            score = extract_score_from_assessment(assessment)
            if score > 0:
                violations_list.append(float(score))

        # 计算统计周期月数
        months_active = 1
        if start_date and end_date:
            # 如果指定了日期范围，计算该范围的月数
            try:
                start = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end = datetime.strptime(end_date + '-01', '%Y-%m-%d')
                months_active = max(1, int((end - start).days / 30) + 1)
            except:
                months_active = 1
        elif start_date:
            # 只指定了开始日期，从开始日期到现在
            try:
                start = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                months_active = max(1, int((datetime.now() - start).days / 30) + 1)
            except:
                months_active = 1
        elif entry_date:
            # 没有日期筛选，使用入职以来的月数
            try:
                entry = datetime.strptime(entry_date, '%Y-%m-%d')
                months_active = max(1, int((datetime.now() - entry).days / 30))
            except:
                months_active = 1

        # 使用双轨评分模型
        safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
        safety_score = safety_result['final_score']
        safety_status_color = safety_result['status_color']
        safety_alert_tag = safety_result['alert_tag']

        # 3. 工作绩效（使用双算法系统，应用日期筛选）
        is_monthly = (start_date == end_date) if start_date and end_date else True

        perf_query = """
            SELECT score, grade, year, month
            FROM performance_records
            WHERE emp_no = ?
        """
        perf_params = [emp_no]

        if start_date:
            perf_query += " AND (year || '-' || printf('%02d', month)) >= ?"
            perf_params.append(start_date)

        if end_date:
            perf_query += " AND (year || '-' || printf('%02d', month)) <= ?"
            perf_params.append(end_date)

        perf_query += " ORDER BY year, month"
        cur.execute(perf_query, perf_params)
        perf_rows = cur.fetchall()

        if perf_rows:
            if is_monthly and len(perf_rows) == 1:
                # 月度快照算法
                score, grade, year, month = perf_rows[0]
                raw_score = float(score) if score else 95
                grade = grade if grade else 'B+'
                perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
                performance_score = perf_result['radar_value']
            else:
                # 周期加权算法（带时间衰减）
                grade_list = [row[1] if row[1] else 'B+' for row in perf_rows]
                grade_dates = [f"{row[2]}-{row[3]:02d}" for row in perf_rows]  # 构建日期列表
                perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
                performance_score = perf_result['radar_value']
        else:
            # 没有绩效数据
            performance_score = 0

        # 4. 学习能力评估（基于综合分的位置+动能算法）
        # 计算当前周期的综合三维分
        current_comprehensive = (
            performance_score * score_weights.get('performance', 0.35) +
            safety_score * score_weights.get('safety', 0.30) +
            training_score * score_weights.get('training', 0.20)
        )

        # 计算上一周期的综合三维分
        is_monthly = (start_date == end_date) if start_date and end_date else True
        previous_comprehensive = 0
        learning_result = None

        if is_monthly and start_date:
            # 月度模式：计算上月同期数据
            try:
                current_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                prev_dt = current_dt.replace(day=1) - timedelta(days=1)
                prev_date = prev_dt.strftime('%Y-%m')

                # 查询上月绩效
                cur.execute("""
                    SELECT score, grade FROM performance_records
                    WHERE emp_no = ? AND (year || '-' || printf('%02d', month)) = ?
                """, [emp_no, prev_date])
                prev_perf_row = cur.fetchone()
                if prev_perf_row:
                    prev_perf_score = calculate_performance_score_monthly(
                        prev_perf_row[1] if prev_perf_row[1] else 'B+',
                        float(prev_perf_row[0]) if prev_perf_row[0] else 95,
                        algo_config
                    )['radar_value']
                else:
                    prev_perf_score = 0

                # 查询上月安全 (FIXED)
                cur.execute("""
                    SELECT assessment
                    FROM safety_inspection_records
                    WHERE inspected_person = ? AND strftime('%Y-%m', inspection_date) = ?
                """, [emp_name, prev_date])
                prev_violations = []
                for safety_row in cur.fetchall():
                    score = extract_score_from_assessment(safety_row[0])
                    if score > 0:
                        prev_violations.append(float(score))
                prev_safety_result = calculate_safety_score_dual_track(prev_violations, months_active=1, config=algo_config)
                prev_safety_score = prev_safety_result['final_score']


                # 查询上月培训
                cur.execute("""
                    SELECT score, is_qualified, is_disqualified, training_date FROM training_records
                    WHERE emp_no = ? AND strftime('%Y-%m', training_date) = ?
                """, [emp_no, prev_date])
                prev_training_rows = cur.fetchall()
                prev_training_result = calculate_training_score_with_penalty(prev_training_rows, duration_days=30, cert_years=cert_years, config=algo_config)
                prev_training_score = prev_training_result['radar_score']  # 修复：使用正确的键名

                # 计算上月综合分
                previous_comprehensive = (
                    prev_perf_score * score_weights.get('performance', 0.35) +
                    prev_safety_score * score_weights.get('safety', 0.30) +
                    prev_training_score * score_weights.get('training', 0.20)
                )

                # 使用月度算法
                learning_result = calculate_learning_ability_monthly(current_comprehensive, previous_comprehensive)
            except Exception as e:
                # 异常情况：使用当前分作为上月分（视为无变化）
                learning_result = calculate_learning_ability_monthly(current_comprehensive, current_comprehensive)
        else:
            # 长周期模式：查询过去N个月的综合分列表（与右侧API一致）
            try:
                # 获取起止月份
                if start_date and end_date:
                    start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                    end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
                else:
                    end_dt = datetime.now()
                    start_dt = end_dt - timedelta(days=365)

                # 构建月份列表
                month_list = []
                current_month = start_dt
                while current_month <= end_dt:
                    month_list.append(current_month.strftime('%Y-%m'))
                    # 移动到下个月
                    current_month = (current_month.replace(day=1) + timedelta(days=32)).replace(day=1)

                # 循环查询每月三维分
                score_list = []
                for month_str in month_list:
                    # 查询该月绩效
                    cur.execute("""
                        SELECT score, grade FROM performance_records
                        WHERE emp_no = ? AND (year || '-' || printf('%02d', month)) = ?
                    """, [emp_no, month_str])
                    month_perf = cur.fetchone()
                    if month_perf:
                        month_perf_score = calculate_performance_score_monthly(
                            month_perf[1] if month_perf[1] else 'B+',
                            float(month_perf[0]) if month_perf[0] else 95,
                            algo_config
                        )['radar_value']
                    else:
                        month_perf_score = 0

                    # 查询该月安全
                    cur.execute("""
                        SELECT assessment FROM safety_inspection_records
                        WHERE inspected_person = ? AND strftime('%Y-%m', inspection_date) = ?
                    """, [emp_name, month_str])
                    month_safety_rows = cur.fetchall()
                    if month_safety_rows:
                        month_violations = []
                        for safety_row in month_safety_rows:  # 修复：避免覆盖外层row变量
                            score = extract_score_from_assessment(safety_row[0])
                            if score > 0:
                                month_violations.append(float(score))
                        month_safety_result = calculate_safety_score_dual_track(month_violations, 1, algo_config)
                        month_safety_score = month_safety_result['final_score']
                    else:
                        month_safety_score = 0

                    # 查询该月培训
                    cur.execute("""
                        SELECT score, is_qualified, is_disqualified, training_date FROM training_records
                        WHERE emp_no = ? AND strftime('%Y-%m', training_date) = ?
                    """, [emp_no, month_str])
                    month_training_rows = cur.fetchall()
                    if month_training_rows:
                        month_training_result = calculate_training_score_with_penalty(
                            month_training_rows,
                            30,  # 单月30天
                            cert_years,
                            algo_config
                        )
                        month_training_score = month_training_result['radar_score']
                    else:
                        month_training_score = 0

                    # 计算该月综合分（使用配置权重）
                    month_comprehensive = (
                        month_perf_score * score_weights.get('performance', 0.35) +
                        month_safety_score * score_weights.get('safety', 0.30) +
                        month_training_score * score_weights.get('training', 0.20)
                    )
                    score_list.append(month_comprehensive)

                # 使用长周期算法
                if len(score_list) >= 2:
                    print(f"DEBUG [api_students_list-员工{emp_no}]: 使用长周期算法，score_list长度={len(score_list)}, current_comprehensive={current_comprehensive:.1f}")
                    learning_result = calculate_learning_ability_longterm(
                        score_list,
                        algo_config,
                        current_three_dim_score=current_comprehensive
                    )
                    print(f"DEBUG [api_students_list-员工{emp_no}]: 学习能力分数={learning_result['learning_score']}")
                else:
                    # 数据不足，使用月度算法
                    print(f"DEBUG [api_students_list-员工{emp_no}]: 数据不足(len={len(score_list)})，降级到月度算法")
                    learning_result = calculate_learning_ability_monthly(
                        current_comprehensive,
                        current_comprehensive
                    )
            except Exception as e:
                # 异常情况：使用当前分
                print(f"ERROR [api_students_list-员工{emp_no}]: 学习能力计算异常 - {type(e).__name__}: {e}")
                import traceback
                traceback.print_exc()
                learning_result = calculate_learning_ability_monthly(
                    current_comprehensive,
                    current_comprehensive
                )

        # 提取学习能力分值
        if learning_result:
            learning_score = learning_result['learning_score']
        else:
            learning_score = 0

        # 5. 稳定性（使用完整算法：资历60% + 表现稳定性40%）
        entry_date = safe_get(row, 'entry_date')
        birth_date = safe_get(row, 'birth_date')
        work_start_date = safe_get(row, 'work_start_date')
        cert_date = safe_get(row, 'certification_date')
        solo_date = safe_get(row, 'solo_driving_date')

        try:
            # 重新定义日期范围以避免作用域问题
            if start_date and end_date:
                start_dt_stability = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end_dt_stability = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            else:
                # 如果没有筛选，使用过去12个月
                end_dt_stability = datetime.now()
                start_dt_stability = end_dt_stability - timedelta(days=365)

            # 构建月份列表
            month_list = []
            current_month = start_dt_stability.replace(day=1)
            while current_month <= end_dt_stability:
                month_list.append(current_month.strftime('%Y-%m'))
                # 移动到下个月
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)

            # 查询每个月的三维分数
            historical_scores = {
                'performance': [],
                'safety': [],
                'training': []
            }

            for month_str in month_list:
                # 查询该月绩效分
                cur.execute("""
                    SELECT score, grade FROM performance_records
                    WHERE emp_no = ? AND (year || '-' || printf('%02d', month)) = ?
                """, [emp_no, month_str])
                month_perf_row = cur.fetchone()
                if month_perf_row:
                    month_perf_score = calculate_performance_score_monthly(
                        month_perf_row[1] if month_perf_row[1] else 'B+',
                        float(month_perf_row[0]) if month_perf_row[0] else 95,
                        algo_config
                    )['radar_value']
                    historical_scores['performance'].append(month_perf_score)

                # 查询该月安全分
                cur.execute("""
                    SELECT assessment, inspection_date
                    FROM safety_inspection_records
                    WHERE inspected_person = ? AND strftime('%Y-%m', inspection_date) = ?
                    ORDER BY inspection_date
                """, [emp_name, month_str])
                month_safety_rows = cur.fetchall()
                if month_safety_rows:
                    violations = []
                    for safety_row in month_safety_rows:  # 修复：避免覆盖外层row变量
                        score = extract_score_from_assessment(safety_row[0])
                        if score > 0:
                            violations.append(float(score))

                    if violations:
                        month_safety_result = calculate_safety_score_dual_track(
                            violations,
                            1,
                            algo_config
                        )
                        historical_scores['safety'].append(month_safety_result['final_score'])

                # 查询该月培训分
                cur.execute("""
                    SELECT score, is_qualified, is_disqualified, training_date FROM training_records
                    WHERE emp_no = ? AND strftime('%Y-%m', training_date) = ?
                """, [emp_no, month_str])
                month_training_rows = cur.fetchall()
                if month_training_rows:
                    month_training_result = calculate_training_score_with_penalty(
                        month_training_rows,
                        30,
                        cert_years,
                        algo_config
                    )
                    historical_scores['training'].append(month_training_result['radar_score'])

            # 调用综合稳定性算法
            print(f"DEBUG [api_students_list-员工{emp_no}]: 稳定性算法参数:")
            print(f"  - birth_date={birth_date}, work_start_date={work_start_date}")
            print(f"  - entry_date={entry_date}, cert_date={cert_date}, solo_date={solo_date}")
            print(f"  - historical_scores: perf={len(historical_scores['performance'])}条, safety={len(historical_scores['safety'])}条, training={len(historical_scores['training'])}条")
            stability_result = calculate_stability_score(
                birth_date=birth_date,
                work_start_date=work_start_date,
                entry_date=entry_date,
                certification_date=cert_date,
                solo_driving_date=solo_date,
                historical_scores=historical_scores if any(historical_scores.values()) else None,
                config=algo_config
            )
            stability_score = stability_result['stability_score']
            print(f"DEBUG [api_students_list-员工{emp_no}]: 稳定性分数={stability_score:.1f}（综合算法）")
            print(f"  - 资历分={stability_result['seniority_score']:.1f}, 波动分={stability_result['volatility_score']:.1f}")
            print(f"  - 波动系数={stability_result['metrics']['volatility']:.2f}")
        except Exception as e:
            # 异常情况：使用简单计算作为降级方案
            print(f"ERROR [api_students_list-员工{emp_no}]: 稳定性算法异常 - {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            if entry_date:
                try:
                    entry = datetime.strptime(entry_date, '%Y-%m-%d')
                    years = (datetime.now() - entry).days / 365
                    stability_score = min(100, years * 33.3)
                    print(f"DEBUG [api_students_list-员工{emp_no}]: 降级到简单算法，稳定性={stability_score:.1f}（入职{years:.1f}年）")
                except:
                    stability_score = 50
                    print(f"DEBUG [api_students_list-员工{emp_no}]: 降级失败，使用默认值50")
            else:
                stability_score = 50
                print(f"DEBUG [api_students_list-员工{emp_no}]: 无入职日期，使用默认值50")

        # 综合评分（加权平均 - 使用配置权重）
        comprehensive_score = round(
            performance_score * score_weights['performance'] +
            safety_score * score_weights['safety'] +
            training_score * score_weights['training'] +
            stability_score * score_weights['stability'] +
            learning_score * score_weights['learning'],
            1
        )

        # 判断是否为关键人员（基于筛选日期范围）（使用配置阈值）
        # 复用已计算的违规数据和月数，避免重复查询
        import math
        violation_count = len(violations_list)
        avg_freq = math.ceil(violation_count / months_active) if months_active > 0 else 0

        is_key_personnel = (comprehensive_score < key_personnel_config['comprehensive_threshold']) or (avg_freq >= key_personnel_config['monthly_violation_threshold'])

        students.append({
            'emp_no': emp_no,
            'name': emp_name,
            'department_name': dept_name,
            'position': safe_get(row, 'position'),
            'comprehensive_score': comprehensive_score,
            'is_key_personnel': bool(is_key_personnel),  # 显式转换为JSON兼容的布尔值
            'safety_status_color': safety_status_color,
            'safety_alert_tag': safety_alert_tag
        })

    # 按综合分升序排序
    students.sort(key=lambda x: x['comprehensive_score'])

    return jsonify(students)


@personnel_bp.route('/api/comprehensive-profile/<emp_no>')
@login_required
def api_comprehensive_profile(emp_no):
    """API: 获取个人综合能力画像（人员+培训+安全+绩效）"""
    from datetime import datetime, timedelta
    from blueprints.safety import extract_score_from_assessment

    # 读取算法配置
    from services.algorithm_config_service import AlgorithmConfigService
    algo_config = AlgorithmConfigService.get_active_config()
    score_weights = algo_config['comprehensive']['score_weights']

    conn = get_db()
    cur = conn.cursor()

    # 1. 获取员工基本信息
    cur.execute("""
        SELECT
            name, department_id, position, education, entry_date,
            birth_date, work_start_date, certification_date, solo_driving_date
        FROM employees
        WHERE emp_no = ?
    """, (emp_no,))
    employee = cur.fetchone()

    if not employee:
        return jsonify({'error': '员工不存在'}), 404

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    emp_name, dept_id, position, education, entry_date, \
        birth_date, work_start_date, cert_date, solo_date = employee

    # 计算各项年限
    working_years = calculate_years_from_date(work_start_date) if work_start_date else None
    tenure_years = calculate_years_from_date(entry_date) if entry_date else None
    cert_years = calculate_years_from_date(cert_date) if cert_date else None
    solo_years = calculate_years_from_date(solo_date) if solo_date else None

    # 获取日期筛选参数（如果有）
    start_date = request.args.get('start_date')  # 格式：YYYY-MM
    end_date = request.args.get('end_date')      # 格式：YYYY-MM

    # DEBUG: 打印接收到的日期参数
    print(f"DEBUG [comprehensive-profile]: 原始参数 - start_date='{start_date}', end_date='{end_date}'")
    print(f"DEBUG [comprehensive-profile]: 参数类型 - start_date type={type(start_date)}, end_date type={type(end_date)}")
    print(f"DEBUG [comprehensive-profile]: 参数布尔值 - bool(start_date)={bool(start_date)}, bool(end_date)={bool(end_date)}")

    # 如果没有指定日期，默认使用当月
    if not start_date and not end_date:
        current_month = datetime.now().strftime('%Y-%m')
        start_date = current_month
        end_date = current_month
        print(f"DEBUG [comprehensive-profile]: 无日期参数，使用默认当月: {current_month}")

    # 2. 培训能力分析（使用高级评分算法 - 包含毒性惩罚和动态年化）
    training_query = """
        SELECT
            score,
            is_qualified,
            is_disqualified,
            training_date
        FROM training_records
        WHERE emp_no = ?
    """
    training_params = [emp_no]

    if start_date:
        training_query += " AND strftime('%Y-%m', training_date) >= ?"
        training_params.append(start_date)

    if end_date:
        training_query += " AND strftime('%Y-%m', training_date) <= ?"
        training_params.append(end_date)

    training_query += " ORDER BY training_date ASC"
    cur.execute(training_query, training_params)
    training_records = cur.fetchall()

    # 计算统计周期天数
    if start_date and end_date and start_date == end_date:
        # 单月统计，按30天计算
        duration_days = 30
    elif start_date and end_date:
        # 多月统计，计算实际天数
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            # 计算到月末
            import calendar
            end_year, end_month = int(end_date.split('-')[0]), int(end_date.split('-')[1])
            last_day = calendar.monthrange(end_year, end_month)[1]
            end_dt = end_dt.replace(day=last_day)
            duration_days = max(1, (end_dt - start_dt).days + 1)
        except:
            duration_days = 30
    else:
        # 默认按30天计算
        duration_days = 30

    # 使用新的评分算法
    training_result = calculate_training_score_with_penalty(training_records, duration_days, cert_years, algo_config)
    training_score = training_result['radar_score']
    training_status_color = training_result['status_color']
    training_alert_tag = training_result['alert_tag']
    training_original_score = training_result['original_score']
    training_penalty_coeff = training_result['penalty_coefficient']
    total_training_count = training_result['stats']['total_ops']
    training_fail_count = training_result['stats']['fail_count']

    # 3. 安全能力分析（使用双轨评分模型，应用日期筛选）
    safety_query = """
        SELECT
            inspection_date,
            assessment,
            inspected_person,
            rectifier
        FROM safety_inspection_records
        WHERE (inspected_person = ? OR rectifier = ?)
    """
    safety_params = [emp_name, emp_name]

    if start_date:
        safety_query += " AND strftime('%Y-%m', inspection_date) >= ?"
        safety_params.append(start_date)

    if end_date:
        safety_query += " AND strftime('%Y-%m', inspection_date) <= ?"
        safety_params.append(end_date)

    safety_query += " ORDER BY inspection_date ASC"
    cur.execute(safety_query, safety_params)

    violations_list = []
    safety_as_inspector = 0
    safety_as_rectifier = 0

    for row in cur.fetchall():
        date, assessment, inspected, rectifier = row
        score = extract_score_from_assessment(assessment)

        if inspected == emp_name and score > 0:
            violations_list.append(float(score))

        if inspected == emp_name:
            safety_as_inspector += 1
        if rectifier == emp_name:
            safety_as_rectifier += 1

    # 计算统计周期月数（使用筛选日期范围的月数）
    months_active = 1
    if start_date and end_date:
        # 如果指定了日期范围，计算该范围的月数
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            months_active = max(1, int((end_dt - start_dt).days / 30) + 1)
        except:
            months_active = 1
    elif start_date:
        # 只指定了开始日期，从开始日期到现在
        try:
            start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            months_active = max(1, int((datetime.now() - start_dt).days / 30) + 1)
        except:
            months_active = 1
    elif entry_date:
        # 没有日期筛选，使用入职以来的月数
        try:
            entry = datetime.strptime(entry_date, '%Y-%m-%d')
            months_active = max(1, int((datetime.now() - entry).days / 30))
        except:
            months_active = 1

    # 使用双轨评分模型
    safety_result = calculate_safety_score_dual_track(violations_list, months_active, algo_config)
    safety_score = safety_result['final_score']
    safety_status_color = safety_result['status_color']
    safety_alert_tag = safety_result['alert_tag']
    safety_violations = len(violations_list)
    safety_total_score = sum(violations_list)

    # 4. 绩效能力分析（使用双算法系统）
    # 判断是月度还是周期（使用前面已经设置的 start_date 和 end_date）
    is_monthly = (start_date == end_date) if start_date and end_date else True
    print(f"DEBUG [comprehensive-profile]: is_monthly={is_monthly}, start_date={start_date}, end_date={end_date}")

    # 构建绩效查询
    perf_query = """
        SELECT score, grade, year, month
        FROM performance_records
        WHERE emp_no = ?
    """
    perf_params = [emp_no]

    if start_date:
        perf_query += " AND (year || '-' || printf('%02d', month)) >= ?"
        perf_params.append(start_date)

    if end_date:
        perf_query += " AND (year || '-' || printf('%02d', month)) <= ?"
        perf_params.append(end_date)

    perf_query += " ORDER BY year, month"
    cur.execute(perf_query, perf_params)
    perf_rows = cur.fetchall()

    if perf_rows:
        if is_monthly and len(perf_rows) == 1:
            # 月度快照算法
            score, grade, year, month = perf_rows[0]
            raw_score = float(score) if score else 95
            grade = grade if grade else 'B+'
            perf_result = calculate_performance_score_monthly(grade, raw_score, algo_config)
            performance_score = perf_result['radar_value']
            performance_status_color = perf_result['status_color']
            performance_alert_tag = perf_result['alert_tag']
            performance_display_label = perf_result['display_label']
            performance_mode = 'MONTHLY'
        else:
            # 周期加权算法（带时间衰减）
            grade_list = [row[1] if row[1] else 'B+' for row in perf_rows]
            grade_dates = [f"{row[2]}-{row[3]:02d}" for row in perf_rows]  # 构建日期列表
            perf_result = calculate_performance_score_period(grade_list, grade_dates, algo_config)
            performance_score = perf_result['radar_value']
            performance_status_color = perf_result['status_color']
            performance_alert_tag = perf_result['alert_tag']
            performance_display_label = perf_result['display_label']
            performance_mode = 'PERIOD'
        performance_count = len(perf_rows)
    else:
        # 没有绩效数据
        performance_score = 0
        performance_count = 0
        performance_status_color = 'GREEN'
        performance_alert_tag = '暂无数据'
        performance_display_label = '暂无数据'
        performance_mode = 'MONTHLY'

    # 5. 学习能力评估（基于综合分的位置+动能算法）
    # 计算当前周期的综合三维分（绩效+安全+培训加权平均）
    current_comprehensive = (
        performance_score * score_weights.get('performance', 0.35) +
        safety_score * score_weights.get('safety', 0.30) +
        training_score * score_weights.get('training', 0.20)
    )

    # 计算上一周期的综合三维分
    previous_comprehensive = 0
    learning_result = None

    # DEBUG: 打印学习能力计算模式
    print(f"DEBUG [学习能力]: is_monthly={is_monthly}, start_date='{start_date}', end_date='{end_date}'")
    print(f"DEBUG [学习能力]: 条件判断 - (is_monthly and start_date) = {is_monthly and start_date}")

    if is_monthly and start_date:
        # 月度模式：计算上月同期数据
        try:
            current_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            prev_dt = current_dt.replace(day=1) - timedelta(days=1)
            prev_date = prev_dt.strftime('%Y-%m')

            # 查询上月绩效
            cur.execute("""
                SELECT score, grade FROM performance_records
                WHERE emp_no = ? AND (year || '-' || printf('%02d', month)) = ?
            """, [emp_no, prev_date])
            prev_perf_row = cur.fetchone()
            if prev_perf_row:
                prev_perf_score = calculate_performance_score_monthly(
                    prev_perf_row[1] if prev_perf_row[1] else 'B+',
                    float(prev_perf_row[0]) if prev_perf_row[0] else 95,
                    algo_config
                )['radar_value']
            else:
                prev_perf_score = 0

            # 查询上月安全 (FIXED)
            cur.execute("""
                SELECT assessment
                FROM safety_inspection_records
                WHERE inspected_person = ? AND strftime('%Y-%m', inspection_date) = ?
            """, [emp_name, prev_date])

            prev_violations = []
            for row in cur.fetchall():
                score = extract_score_from_assessment(row[0])
                if score > 0:
                    prev_violations.append(float(score))

            # 月度模式，周期为1个月
            prev_safety_result = calculate_safety_score_dual_track(prev_violations, months_active=1, config=algo_config)
            prev_safety_score = prev_safety_result['final_score']

            # 查询上月培训
            cur.execute("""
                SELECT score, is_qualified, is_disqualified, training_date FROM training_records
                WHERE emp_no = ? AND strftime('%Y-%m', training_date) = ?
            """, [emp_no, prev_date])
            prev_training_rows = cur.fetchall()
            # 月度模式，周期30天
            prev_training_result = calculate_training_score_with_penalty(prev_training_rows, duration_days=30, cert_years=cert_years, config=algo_config)
            prev_training_score = prev_training_result['radar_score']  # 修复：使用正确的键名


            # 计算上月综合分
            previous_comprehensive = (
                prev_perf_score * score_weights.get('performance', 0.35) +
                prev_safety_score * score_weights.get('safety', 0.30) +
                prev_training_score * score_weights.get('training', 0.20)
            )

            # 使用月度算法
            learning_result = calculate_learning_ability_monthly(current_comprehensive, previous_comprehensive)
        except Exception as e:
            # 异常情况：使用当前分作为上月分（视为无变化）
            learning_result = calculate_learning_ability_monthly(current_comprehensive, current_comprehensive)
    else:
        # 长周期模式：查询过去12个月的综合分列表
        try:
            # 获取起止月份
            if start_date and end_date:
                start_dt = datetime.strptime(start_date + '-01', '%Y-%m-%d')
                end_dt = datetime.strptime(end_date + '-01', '%Y-%m-%d')
            else:
                end_dt = datetime.now()
                start_dt = end_dt - timedelta(days=365)

            # 构建月份列表
            month_list = []
            current_month = start_dt
            while current_month <= end_dt:
                month_list.append(current_month.strftime('%Y-%m'))
                current_month = current_month + timedelta(days=32)
                current_month = current_month.replace(day=1)

            print(f"DEBUG: 构建了 {len(month_list)} 个月份: {month_list}")

            # 查询每个月的三维分数并计算综合分
            score_list = []
            for month_str in month_list:
                print(f"DEBUG: 处理月份 {month_str}")
                # 绩效
                cur.execute("""
                    SELECT score, grade FROM performance_records
                    WHERE emp_no = ? AND (year || '-' || printf('%02d', month)) = ?
                """, [emp_no, month_str])
                month_perf_row = cur.fetchone()
                if month_perf_row:
                    month_perf_score = calculate_performance_score_monthly(
                        month_perf_row[1] if month_perf_row[1] else 'B+',
                        float(month_perf_row[0]) if month_perf_row[0] else 95,
                        algo_config
                    )['radar_value']
                    print(f"  - 绩效: {month_perf_score} (grade={month_perf_row[1]}, score={month_perf_row[0]})")
                else:
                    month_perf_score = 0
                    print(f"  - 绩效: 无数据")

                # 安全
                cur.execute("""
                    SELECT assessment, inspection_date
                    FROM safety_inspection_records
                    WHERE inspected_person = ? AND strftime('%Y-%m', inspection_date) = ?
                    ORDER BY inspection_date
                """, [emp_name, month_str])
                month_safety_rows = cur.fetchall()
                if month_safety_rows:
                    # 提取扣分数值
                    violations = []
                    for row in month_safety_rows:
                        score = extract_score_from_assessment(row[0])
                        if score > 0:
                            violations.append(float(score))

                    if violations:
                        month_safety_result = calculate_safety_score_dual_track(
                            violations,
                            1,
                            algo_config
                        )
                        month_safety_score = month_safety_result['final_score']
                        print(f"  - 安全: {month_safety_score} ({len(violations)}条违规)")
                    else:
                        month_safety_score = 0
                        print(f"  - 安全: 0 (有记录但无扣分)")
                else:
                    month_safety_score = 0
                    print(f"  - 安全: 无数据")

                # 培训
                cur.execute("""
                    SELECT score, is_qualified, is_disqualified, training_date FROM training_records
                    WHERE emp_no = ? AND strftime('%Y-%m', training_date) = ?
                """, [emp_no, month_str])
                month_training_rows = cur.fetchall()
                if month_training_rows:
                    month_training_result = calculate_training_score_with_penalty(
                        month_training_rows,
                        30,  # 单月30天
                        cert_years,
                        algo_config
                    )
                    month_training_score = month_training_result['radar_score']  # 修复：使用正确的键名
                    print(f"  - 培训: {month_training_score} ({len(month_training_rows)}条记录)")
                else:
                    month_training_score = 0
                    print(f"  - 培训: 无数据")

                # 计算该月综合分（使用配置权重）
                month_comprehensive = (
                    month_perf_score * score_weights.get('performance', 0.35) +
                    month_safety_score * score_weights.get('safety', 0.30) +
                    month_training_score * score_weights.get('training', 0.20)
                )
                print(f"  → 综合分: {month_comprehensive:.2f}")
                score_list.append(month_comprehensive)

            # 使用长周期算法
            print(f"DEBUG: score_list 长度 = {len(score_list)}, 内容前3项 = {score_list[:3]}")
            if len(score_list) >= 2:
                print(f"DEBUG: 使用长周期算法，score_list 完整内容 = {score_list}")
                learning_result = calculate_learning_ability_longterm(
                    score_list,
                    algo_config,
                    current_three_dim_score=current_comprehensive  # 传入当前三维综合分
                )
                print(f"DEBUG: 长周期算法返回 = {learning_result}")
            else:
                # 数据不足，使用月度算法
                print(f"DEBUG: 数据不足 (len={len(score_list)})，使用月度算法")
                learning_result = calculate_learning_ability_monthly(current_comprehensive, current_comprehensive)
        except Exception as e:
            # 异常情况：使用当前分
            print(f"ERROR: 学习能力计算异常 - {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            learning_result = calculate_learning_ability_monthly(current_comprehensive, current_comprehensive)

    # 提取学习能力分值和详情
    if learning_result:
        learning_score = learning_result['learning_score']
        learning_status_color = learning_result['status_color']
        learning_alert_tag = learning_result['alert_tag']
        learning_tier = learning_result['tier']
        learning_delta = learning_result.get('delta', 0)
        learning_slope = learning_result.get('slope', 0)
    else:
        learning_score = 0
        learning_status_color = 'GRAY'
        learning_alert_tag = '暂无数据'
        learning_tier = '无数据'
        learning_delta = 0
        learning_slope = 0

    # 6. 稳定性评估（综合算法：资历60% + 表现稳定性40%）
    # 查询用户筛选日期范围内的历史分数用于波动度计算
    try:
        from datetime import datetime, timedelta
        import calendar

        # 构建用户筛选日期范围的月份列表（与左侧API一致）
        if start_date and end_date:
            start_dt_stability = datetime.strptime(start_date + '-01', '%Y-%m-%d')
            end_dt_stability = datetime.strptime(end_date + '-01', '%Y-%m-%d')
        else:
            # 如果没有筛选，使用过去12个月
            end_dt_stability = datetime.now()
            start_dt_stability = end_dt_stability - timedelta(days=365)

        month_list = []
        current_month = start_dt_stability.replace(day=1)
        while current_month <= end_dt_stability:
            month_list.append(current_month.strftime('%Y-%m'))
            # 移动到下个月
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year + 1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)

        # 查询每个月的三维分数
        historical_scores = {
            'performance': [],
            'safety': [],
            'training': []
        }

        for month_str in month_list:
            # 查询该月绩效分
            cur.execute("""
                SELECT score, grade FROM performance_records
                WHERE emp_no = ? AND (year || '-' || printf('%02d', month)) = ?
            """, [emp_no, month_str])
            month_perf_row = cur.fetchone()
            if month_perf_row:
                month_perf_score = calculate_performance_score_monthly(
                    month_perf_row[1] if month_perf_row[1] else 'B+',
                    float(month_perf_row[0]) if month_perf_row[0] else 95,
                    algo_config
                )['radar_value']
                historical_scores['performance'].append(month_perf_score)

            # 查询该月安全分
            cur.execute("""
                SELECT assessment, inspection_date
                FROM safety_inspection_records
                WHERE inspected_person = ? AND strftime('%Y-%m', inspection_date) = ?
                ORDER BY inspection_date
            """, [emp_name, month_str])
            month_safety_rows = cur.fetchall()
            if month_safety_rows:
                # 提取扣分值
                violations = []
                for row in month_safety_rows:
                    score = extract_score_from_assessment(row[0])
                    if score > 0:
                        violations.append(float(score))

                if violations:
                    month_safety_result = calculate_safety_score_dual_track(
                        violations,
                        1,  # 单月
                        algo_config
                    )
                    historical_scores['safety'].append(month_safety_result['final_score'])

            # 查询该月培训分
            cur.execute("""
                SELECT score, is_qualified, is_disqualified, training_date FROM training_records
                WHERE emp_no = ? AND strftime('%Y-%m', training_date) = ?
            """, [emp_no, month_str])
            month_training_rows = cur.fetchall()
            if month_training_rows:
                month_training_result = calculate_training_score_with_penalty(
                    month_training_rows,
                    30,  # 单月30天
                    cert_years,
                    algo_config
                )
                historical_scores['training'].append(month_training_result['radar_score'])

        # 调用综合稳定性算法
        print(f"DEBUG [comprehensive-profile-员工{emp_no}]: 稳定性算法参数:")
        print(f"  - birth_date={birth_date}, work_start_date={work_start_date}")
        print(f"  - entry_date={entry_date}, cert_date={cert_date}, solo_date={solo_date}")
        print(f"  - historical_scores: perf={len(historical_scores['performance'])}条, safety={len(historical_scores['safety'])}条, training={len(historical_scores['training'])}条")
        stability_result = calculate_stability_score(
            birth_date=birth_date,
            work_start_date=work_start_date,
            entry_date=entry_date,
            certification_date=cert_date,
            solo_driving_date=solo_date,
            historical_scores=historical_scores if any(historical_scores.values()) else None,
            config=algo_config
        )
        stability_score = stability_result['stability_score']
        print(f"DEBUG [comprehensive-profile-员工{emp_no}]: 稳定性分数={stability_score:.1f}（综合算法）")

    except Exception as e:
        # 异常情况：使用简单计算作为降级方案
        print(f"稳定性算法异常: {e}")
        import traceback
        traceback.print_exc()
        if entry_date:
            try:
                entry = datetime.strptime(entry_date, '%Y-%m-%d')
                years = (datetime.now() - entry).days / 365
                stability_score = min(100, years * 33.3)
            except:
                stability_score = 50
        else:
            stability_score = 50

    # 7. 计算综合能力分数（加权平均 - 使用配置权重）
    comprehensive_score = round(
        performance_score * score_weights['performance'] +
        safety_score * score_weights['safety'] +
        training_score * score_weights['training'] +
        stability_score * score_weights['stability'] +
        learning_score * score_weights['learning'],
        1
    )

    return jsonify({
        'employee': {
            'emp_no': emp_no,
            'name': emp_name,
            'position': position,
            'education': education,
            'entry_date': entry_date
        },
        'scores': {
            'comprehensive': round(comprehensive_score, 1),
            'training': round(training_score, 1),
            'safety': round(safety_score, 1),
            'performance': round(performance_score, 1),
            'learning': round(learning_score, 1),
            'stability': round(stability_score, 1)
        },
        'personnel_details': {
            'working_years': round(working_years, 1) if working_years else None,
            'tenure_years': round(tenure_years, 1) if tenure_years else None,
            'certification_years': round(cert_years, 1) if cert_years else None,
            'solo_driving_years': round(solo_years, 1) if solo_years else None,
            'education': education
        },
        'safety_details': {
            'violations': safety_violations,
            'total_deduction': safety_total_score,
            'as_inspector': safety_as_inspector,
            'as_rectifier': safety_as_rectifier,
            'status_color': safety_status_color,
            'alert_tag': safety_alert_tag,
            'score_a': safety_result['score_a'],
            'score_b': safety_result['score_b'],
            'avg_freq': safety_result['avg_freq']
        },
        'statistics': {
            'total_trainings': total_training_count,
            'avg_training_score': training_score,
            'recent_trainings': len(training_records) if training_records else 0
        },
        'training_details': {
            'radar_score': training_score,
            'original_score': training_original_score,
            'penalty_coefficient': training_penalty_coeff,
            'status_color': training_status_color,
            'alert_tag': training_alert_tag,
            'total_ops': total_training_count,
            'fail_count': training_fail_count,
            'duration_days': duration_days
        },
        'performance_details': {
            'recent_avg': performance_score,
            'range': f'{"当月" if is_monthly else "统计周期"}',
            'count': performance_count,
            'status_color': performance_status_color,
            'alert_tag': performance_alert_tag,
            'display_label': performance_display_label,
            'mode': performance_mode
        },
        'learning_details': {
            'learning_score': round(learning_score, 1),
            'status_color': learning_status_color,
            'alert_tag': learning_alert_tag,
            'tier': learning_tier,
            'delta': round(learning_delta, 1) if learning_delta else 0,
            'slope': round(learning_slope, 3) if learning_slope else 0,
            'current_comprehensive': round(current_comprehensive, 1),
            'previous_comprehensive': round(previous_comprehensive, 1) if previous_comprehensive else 0
        }
    })


@personnel_bp.route('/api/student-detail/<emp_no>')
@login_required
def api_student_detail(emp_no):
    """API: 获取学员详细数据（培训雷达图数据：按项目分类的平均分）"""
    conn = get_db()
    cur = conn.cursor()

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # 获取时间筛选参数
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    # 构建时间筛选条件
    time_filter = ""
    time_params = [emp_no]
    if year and month:
        time_filter = " AND strftime('%Y', training_date) = ? AND strftime('%m', training_date) = ?"
        time_params.extend([str(year), str(month).zfill(2)])
    elif year:
        time_filter = " AND strftime('%Y', training_date) = ?"
        time_params.append(str(year))

    # 查询该学员各项目分类的平均分
    query = f"""
        SELECT
            c.name as category_name,
            ROUND(AVG(tr.score), 1) as avg_score,
            COUNT(*) as count
        FROM training_records tr
        LEFT JOIN training_projects p ON tr.project_id = p.id
        LEFT JOIN training_project_categories c ON p.category_id = c.id
        WHERE tr.emp_no = ? AND c.name IS NOT NULL{time_filter}
        GROUP BY c.id, c.name
        ORDER BY c.display_order ASC
    """
    cur.execute(query, time_params)
    student_data = {}
    for row in cur.fetchall():
        student_data[row[0]] = {
            'avg_score': row[1],
            'count': row[2]
        }

    # 查询团队平均（基于权限过滤的可见员工）
    accessible_dept_ids = get_accessible_department_ids()
    if not accessible_dept_ids:
        return jsonify({
            'student_data': student_data,
            'team_data': {},
            'categories': sorted(list(student_data.keys()))
        })

    placeholders = ','.join('?' * len(accessible_dept_ids))

    # 构建团队查询的时间筛选
    team_time_filter = ""
    team_time_params = accessible_dept_ids.copy()
    if year and month:
        team_time_filter = " AND strftime('%Y', tr.training_date) = ? AND strftime('%m', tr.training_date) = ?"
        team_time_params.extend([str(year), str(month).zfill(2)])
    elif year:
        team_time_filter = " AND strftime('%Y', tr.training_date) = ?"
        team_time_params.append(str(year))

    query = f"""
        SELECT
            c.name as category_name,
            ROUND(AVG(tr.score), 1) as avg_score
        FROM training_records tr
        LEFT JOIN employees e ON tr.emp_no = e.emp_no
        LEFT JOIN training_projects p ON tr.project_id = p.id
        LEFT JOIN training_project_categories c ON p.category_id = c.id
        WHERE (e.department_id IN ({placeholders}) OR e.emp_no IS NULL)
            AND c.name IS NOT NULL{team_time_filter}
        GROUP BY c.id, c.name
        ORDER BY c.display_order ASC
    """
    cur.execute(query, team_time_params)
    team_data = {}
    for row in cur.fetchall():
        team_data[row[0]] = row[1]

    # 合并所有分类
    all_categories = set(student_data.keys()) | set(team_data.keys())

    return jsonify({
        'student_data': student_data,
        'team_data': team_data,
        'categories': sorted(list(all_categories))
    })


@personnel_bp.route('/api/student-growth/<emp_no>')
@login_required
def api_student_growth(emp_no):
    """API: 获取学员成长趋势数据（按时间的平均分变化）"""
    conn = get_db()
    cur = conn.cursor()

    # 验证权限
    if not validate_employee_access(emp_no):
        return jsonify({'error': '无权限查看此员工'}), 403

    # 查询该学员按月份的平均分趋势
    query = """
        SELECT
            strftime('%Y-%m', training_date) as month,
            ROUND(AVG(score), 1) as avg_score,
            COUNT(*) as count
        FROM training_records
        WHERE emp_no = ?
        GROUP BY month
        ORDER BY month ASC
    """
    cur.execute(query, (emp_no,))

    growth_data = []
    for row in cur.fetchall():
        growth_data.append({
            'month': row[0],
            'avg_score': row[1],
            'count': row[2]
        })

    return jsonify(growth_data)
