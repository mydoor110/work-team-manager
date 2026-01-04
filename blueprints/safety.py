#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
安全管理模块
负责安全检查数据管理、记录查询、分析统计等功能
注意: 使用created_by字段设计，便于后续权限改造
"""
import os
import re
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from config.settings import APP_TITLE, EXPORT_DIR, UPLOAD_DIR
from models.database import get_db
from .decorators import login_required, role_required, manager_required
from .helpers import require_user_id, get_accessible_department_ids, build_department_filter, parse_date_filters, build_date_filter_sql, log_import_operation

# 创建 Blueprint
safety_bp = Blueprint('safety', __name__, url_prefix='/safety')


@safety_bp.route('/')
@login_required
def index():
    """安全管理主控制台"""
    feature_cards = [
        {
            "title": "安全总览",
            "description": "查看安全检查记录情况并导出数据。",
            "endpoint": "safety.records",
        },
        {
            "title": "上传Excel",
            "description": "导入安全检查Excel，自动解析检查记录并保存。",
            "endpoint": "safety.upload",
        },
        {
            "title": "安全统计",
            "description": "查看安全统计分析，包含图表可视化。",
            "endpoint": "safety.analytics",
        },
    ]
    return render_template(
        "safety_dashboard.html",
        title=f"安全管理 | {APP_TITLE}",
        feature_cards=feature_cards,
    )


@safety_bp.route('/upload', methods=['GET'])
@login_required
def upload():
    """安全数据上传主页面"""
    return render_template(
        "safety_upload.html",
        title=f"上传安全数据 | {APP_TITLE}",
    )


@safety_bp.route('/upload/inspection', methods=['GET', 'POST'])
@manager_required
def upload_inspection():
    """上传并导入安全检查Excel文件"""
    if request.method == 'POST':
        file_obj = request.files.get("file")
        if not file_obj or file_obj.filename == "":
            flash("请选择要上传的安全检查文件。", "warning")
            return redirect(url_for("safety.upload_inspection"))

        filename = secure_filename(file_obj.filename)

        # 支持 .xlsx 文件
        if not filename.lower().endswith(('.xlsx', '.xls')):
            flash("仅支持 .xlsx 或 .xls 格式", "warning")
            return redirect(url_for("safety.upload_inspection"))

        uid = require_user_id()
        conn = get_db()
        cur = conn.cursor()

        # ====== 方案3: 检查文件是否已导入 ======
        cur.execute("""
            SELECT COUNT(*) FROM safety_inspection_records
            WHERE source_file = ?
        """, (filename,))

        if cur.fetchone()[0] > 0:
            flash(f'文件 "{filename}" 已导入过，请勿重复导入！', 'warning')
            return redirect(url_for("safety.upload_inspection"))

        # 获取当前用户可访问的部门ID列表（用于权限验证）
        accessible_dept_ids = get_accessible_department_ids()

        total_imported = 0
        total_skipped = 0
        total_skipped_no_permission = 0
        errors = []

        try:
            # 读取 Excel 文件
            wb = load_workbook(file_obj, read_only=True, data_only=True)
            sheet = wb.active

            # 找到表头行（假设第1行是表头）
            header_row_idx = 0
            header_values = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                # 检查是否包含关键列
                if any('类别' in str(cell or '') for cell in row):
                    header_row_idx = idx
                    header_values = row
                    break

            if not header_values:
                flash("无法识别Excel表头，请检查文件格式", "warning")
                return redirect(url_for("safety.upload_inspection"))

            # 解析表头 - 使用更精确的匹配逻辑
            col_map = {}
            for idx, h in enumerate(header_values):
                h_str = str(h or "").strip()

                # 精确匹配优先,避免误匹配
                if h_str == '类别' or ('类别' in h_str and len(h_str) <= 4):
                    col_map['category'] = idx
                elif h_str in ['检查日期', '日期'] and 'deadline' not in col_map:
                    # 确保匹配的是检查日期,而不是整改期限
                    col_map['date'] = idx
                elif h_str == '地点' or ('地点' in h_str and '责任' not in h_str):
                    col_map['location'] = idx
                elif '存在隐患和问题' in h_str or ('隐患' in h_str and '问题' in h_str):
                    col_map['hazard_description'] = idx
                elif ('整改措施' in h_str and '整改人' not in h_str and '整改情况' not in h_str and '整改期限' not in h_str):
                    # 匹配"整改措施"或"整改措施及其它意见",但排除其他包含"整改"的列
                    col_map['corrective_measures'] = idx
                elif '整改期限' in h_str or (h_str == '期限' or ('期限' in h_str and '整改' in h_str)):
                    col_map['deadline'] = idx
                elif h_str == '被检查人' or ('被检查' in h_str and '人' in h_str):
                    col_map['inspected_person'] = idx
                elif h_str == '责任车队' or ('责任' in h_str and '车队' in h_str):
                    col_map['responsible_team'] = idx
                elif h_str == '考核情况' or ('考核' in h_str and '情况' in h_str):
                    col_map['assessment'] = idx
                elif h_str == '整改情况' or (h_str.startswith('整改') and '情况' in h_str and '措施' not in h_str):
                    col_map['rectification_status'] = idx
                elif h_str == '整改人' or (h_str.endswith('整改人')):
                    col_map['rectifier'] = idx
                elif h_str == '作业类型' or ('作业' in h_str and '类型' in h_str):
                    col_map['work_type'] = idx
                elif h_str == '责任点位' or ('责任' in h_str and '点位' in h_str):
                    col_map['responsibility_location'] = idx
                elif h_str == '检查项目' or ('检查' in h_str and '项目' in h_str):
                    col_map['inspection_item'] = idx

            if 'category' not in col_map or 'date' not in col_map:
                # 提供详细的错误信息,帮助用户排查问题
                detected_columns = [f"{h}" for h in header_values if h]
                flash(f"缺少必要列（类别、日期）。检测到的列: {', '.join(detected_columns)}", "warning")
                return redirect(url_for("safety.upload_inspection"))

            # 记录列映射关系用于调试
            import logging
            logging.basicConfig(level=logging.INFO)
            logger = logging.getLogger(__name__)
            logger.info(f"安全检查Excel列映射: {col_map}")
            logger.info(f"表头: {header_values}")

            # ====== 第一阶段：收集所有数据并检查重名 ======
            records_to_import = []
            inspected_persons = set()  # 收集所有被检查人

            # 处理数据行
            for row_idx, row_values in enumerate(sheet.iter_rows(min_row=header_row_idx + 2, values_only=True)):
                # 跳过空行
                if all(not str(v or "").strip() for v in row_values):
                    continue

                def get_val(key):
                    idx = col_map.get(key)
                    if idx is not None and idx < len(row_values):
                        return row_values[idx]
                    return None

                category = str(get_val('category') or "").strip()
                date_val = get_val('date')

                if not category:
                    total_skipped += 1
                    continue

                # 转换日期
                inspection_date = None
                if isinstance(date_val, datetime):
                    inspection_date = date_val.strftime('%Y-%m-%d')
                elif isinstance(date_val, (int, float)):
                    # Excel日期数字转换
                    try:
                        base_date = datetime(1899, 12, 30)
                        converted_date = base_date + timedelta(days=date_val)
                        inspection_date = converted_date.strftime('%Y-%m-%d')
                    except:
                        inspection_date = None
                elif isinstance(date_val, str):
                    inspection_date = str(date_val).strip()

                if not inspection_date:
                    total_skipped += 1
                    continue

                # 转换整改期限日期
                deadline_val = get_val('deadline')
                deadline_date = None
                if isinstance(deadline_val, datetime):
                    deadline_date = deadline_val.strftime('%Y-%m-%d')
                elif isinstance(deadline_val, (int, float)):
                    try:
                        base_date = datetime(1899, 12, 30)
                        converted_date = base_date + timedelta(days=deadline_val)
                        deadline_date = converted_date.strftime('%Y-%m-%d')
                    except:
                        deadline_date = None
                elif isinstance(deadline_val, str):
                    deadline_date = str(deadline_val).strip()

                # 获取被检查人
                inspected_person = str(get_val('inspected_person') or "").strip()
                if inspected_person:
                    inspected_persons.add(inspected_person)

                # 收集记录数据
                record_data = {
                    'category': category,
                    'inspection_date': inspection_date,
                    'location': str(get_val('location') or ""),
                    'hazard_description': str(get_val('hazard_description') or ""),
                    'corrective_measures': str(get_val('corrective_measures') or ""),
                    'deadline_date': deadline_date,
                    'inspected_person': inspected_person,
                    'responsible_team': str(get_val('responsible_team') or ""),
                    'assessment': str(get_val('assessment') or ""),
                    'rectification_status': str(get_val('rectification_status') or ""),
                    'rectifier': str(get_val('rectifier') or ""),
                    'work_type': str(get_val('work_type') or ""),
                    'responsibility_location': str(get_val('responsibility_location') or ""),
                    'inspection_item': str(get_val('inspection_item') or ""),
                }
                records_to_import.append(record_data)

            # ====== 第二阶段：检查重名人员 ======
            duplicate_names = {}  # {name: [emp_no1, emp_no2, ...]}

            for person_name in inspected_persons:
                if not person_name:
                    continue

                # 查询该姓名对应的所有工号
                cur.execute("""
                    SELECT emp_no, name, department_id
                    FROM employees
                    WHERE name = ?
                    ORDER BY emp_no
                """, (person_name,))

                matching_employees = cur.fetchall()

                if len(matching_employees) > 1:
                    # 有重名
                    duplicate_names[person_name] = [
                        {'emp_no': row[0], 'name': row[1], 'dept_id': row[2]}
                        for row in matching_employees
                    ]

            # ====== 第三阶段：如果有重名，使用临时文件存储（避免session过大）======
            if duplicate_names:
                import json
                from pathlib import Path

                # 创建临时文件存储待导入数据
                temp_dir = Path(UPLOAD_DIR) / 'temp_imports'
                temp_dir.mkdir(exist_ok=True)

                # 生成唯一的临时文件名
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                temp_filename = f"pending_safety_{uid}_{timestamp}.json"
                temp_filepath = temp_dir / temp_filename

                # 将数据序列化到临时文件
                temp_data = {
                    'filename': filename,
                    'records': records_to_import,
                    'duplicate_names': duplicate_names,
                    'uid': uid,
                    'created_at': timestamp
                }

                with open(temp_filepath, 'w', encoding='utf-8') as f:
                    json.dump(temp_data, f, ensure_ascii=False, indent=2)

                # session只存储临时文件路径（<100字节）
                session['pending_safety_file'] = temp_filename

                return redirect(url_for('safety.confirm_duplicates'))

            # ====== 第四阶段：无重名，直接导入（去重检查 + 权限验证）======
            for record_data in records_to_import:
                # 权限验证：检查被检查人所属部门是否可访问
                inspected_person = record_data.get('inspected_person', '')
                has_permission = True  # 默认有权限（如果没有被检查人）

                if inspected_person:
                    # 查询被检查人的部门
                    cur.execute("""
                        SELECT department_id FROM employees
                        WHERE name = ?
                        LIMIT 1
                    """, (inspected_person,))
                    emp_row = cur.fetchone()

                    if emp_row and emp_row[0]:
                        emp_dept_id = emp_row[0]
                        # 检查是否有权限访问该部门
                        if emp_dept_id not in accessible_dept_ids:
                            has_permission = False
                            total_skipped_no_permission += 1
                            continue

                # 检查是否已存在完全相同的记录（基于关键字段）
                cur.execute("""
                    SELECT COUNT(*) FROM safety_inspection_records
                    WHERE category = ?
                    AND inspection_date = ?
                    AND location = ?
                    AND hazard_description = ?
                """, (
                    record_data['category'],
                    record_data['inspection_date'],
                    record_data['location'],
                    record_data['hazard_description']
                ))

                if cur.fetchone()[0] > 0:
                    # 已存在相同记录，跳过
                    total_skipped += 1
                    continue

                # 插入新记录
                cur.execute("""
                    INSERT INTO safety_inspection_records(
                        category, inspection_date, location, hazard_description,
                        corrective_measures, deadline_date, inspected_person,
                        responsible_team, assessment, rectification_status,
                        rectifier, work_type, responsibility_location,
                        inspection_item, created_by, source_file
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    record_data['category'], record_data['inspection_date'],
                    record_data['location'], record_data['hazard_description'],
                    record_data['corrective_measures'], record_data['deadline_date'],
                    record_data['inspected_person'], record_data['responsible_team'],
                    record_data['assessment'], record_data['rectification_status'],
                    record_data['rectifier'], record_data['work_type'],
                    record_data['responsibility_location'], record_data['inspection_item'],
                    uid, filename
                ))
                total_imported += 1

            conn.commit()

            # 显示结果和记录日志
            total_rows = total_imported + total_skipped + total_skipped_no_permission
            msg_parts = []

            if total_imported > 0:
                msg_parts.append(f"成功导入 {total_imported} 条安全检查记录")
            if total_skipped > 0:
                msg_parts.append(f"{total_skipped} 条重复/空记录被跳过")
            if total_skipped_no_permission > 0:
                msg_parts.append(f"{total_skipped_no_permission} 条记录因无权限被跳过")

            if msg_parts:
                flash("、".join(msg_parts), "success" if total_imported > 0 else "warning")

            # 记录导入操作日志
            log_import_operation(
                module='safety',
                operation='import',
                file_name=filename,
                total_rows=total_rows,
                success_rows=total_imported,
                failed_rows=0,
                skipped_rows=total_skipped + total_skipped_no_permission,
                import_details={
                    'imported': total_imported,
                    'skipped_duplicate': total_skipped,
                    'skipped_no_permission': total_skipped_no_permission,
                    'accessible_departments': len(accessible_dept_ids)
                }
            )

        except Exception as e:
            error_msg = f"处理错误: {str(e)}"
            flash(error_msg, "danger")
            conn.rollback()

            # 记录失败的导入操作
            log_import_operation(
                module='safety',
                operation='import',
                file_name=filename,
                total_rows=0,
                success_rows=0,
                failed_rows=0,
                skipped_rows=0,
                error_message=error_msg
            )

        return redirect(url_for("safety.records"))

    return render_template(
        "safety_upload.html",
        title=f"上传安全检查数据 | {APP_TITLE}",
    )


@safety_bp.route('/upload/confirm-duplicates', methods=['GET', 'POST'])
@manager_required
def confirm_duplicates():
    """处理重名人员的工号选择"""
    import json
    from pathlib import Path

    # 从临时文件加载数据的辅助函数
    def load_temp_data():
        temp_filename = session.get('pending_safety_file')
        if not temp_filename:
            return None

        temp_dir = Path(UPLOAD_DIR) / 'temp_imports'
        temp_filepath = temp_dir / temp_filename

        if not temp_filepath.exists():
            return None

        try:
            with open(temp_filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"加载临时文件失败: {e}")
            return None

    # 清理临时文件的辅助函数
    def cleanup_temp_file():
        temp_filename = session.get('pending_safety_file')
        if temp_filename:
            temp_dir = Path(UPLOAD_DIR) / 'temp_imports'
            temp_filepath = temp_dir / temp_filename
            try:
                if temp_filepath.exists():
                    temp_filepath.unlink()
            except Exception as e:
                print(f"删除临时文件失败: {e}")
            session.pop('pending_safety_file', None)

    pending_data = load_temp_data()

    if not pending_data:
        flash("没有待处理的导入数据", "warning")
        cleanup_temp_file()
        return redirect(url_for("safety.upload_inspection"))

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'cancel':
            # 用户取消导入
            cleanup_temp_file()
            flash("已取消导入", "info")
            return redirect(url_for("safety.upload_inspection"))

        # 用户确认导入，获取选择的工号
        emp_no_selections = {}  # {person_name: selected_emp_no}

        for person_name in pending_data['duplicate_names'].keys():
            selected_emp_no = request.form.get(f'emp_no_{person_name}')
            if selected_emp_no:
                emp_no_selections[person_name] = selected_emp_no
            else:
                flash(f"请为 {person_name} 选择工号", "warning")
                return redirect(url_for('safety.confirm_duplicates'))

        # 执行导入
        conn = get_db()
        cur = conn.cursor()
        uid = pending_data['uid']
        filename = pending_data['filename']
        records = pending_data['records']

        # 获取当前用户可访问的部门ID列表（用于权限验证）
        accessible_dept_ids = get_accessible_department_ids()

        total_imported = 0
        total_skipped = 0
        total_skipped_no_permission = 0

        try:
            for record_data in records:
                inspected_person = record_data.get('inspected_person', '')

                # 如果这个人员有重名，使用用户选择的工号进行权限验证
                if inspected_person in emp_no_selections:
                    selected_emp_no = emp_no_selections[inspected_person]
                    # 查询选中员工的部门，验证权限
                    cur.execute("""
                        SELECT department_id FROM employees
                        WHERE emp_no = ?
                    """, (selected_emp_no,))
                    emp_row = cur.fetchone()

                    if emp_row and emp_row[0]:
                        emp_dept_id = emp_row[0]
                        # 检查是否有权限访问该部门
                        if emp_dept_id not in accessible_dept_ids:
                            total_skipped_no_permission += 1
                            continue
                elif inspected_person:
                    # 无重名情况，按名字查询部门验证权限
                    cur.execute("""
                        SELECT department_id FROM employees
                        WHERE name = ?
                        LIMIT 1
                    """, (inspected_person,))
                    emp_row = cur.fetchone()

                    if emp_row and emp_row[0]:
                        emp_dept_id = emp_row[0]
                        if emp_dept_id not in accessible_dept_ids:
                            total_skipped_no_permission += 1
                            continue

                # 检查是否已存在完全相同的记录
                cur.execute("""
                    SELECT COUNT(*) FROM safety_inspection_records
                    WHERE category = ?
                    AND inspection_date = ?
                    AND location = ?
                    AND hazard_description = ?
                """, (
                    record_data['category'],
                    record_data['inspection_date'],
                    record_data['location'],
                    record_data['hazard_description']
                ))

                if cur.fetchone()[0] > 0:
                    total_skipped += 1
                    continue

                # 插入记录
                cur.execute("""
                    INSERT INTO safety_inspection_records(
                        category, inspection_date, location, hazard_description,
                        corrective_measures, deadline_date, inspected_person,
                        responsible_team, assessment, rectification_status,
                        rectifier, work_type, responsibility_location,
                        inspection_item, created_by, source_file
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    record_data['category'], record_data['inspection_date'],
                    record_data['location'], record_data['hazard_description'],
                    record_data['corrective_measures'], record_data['deadline_date'],
                    record_data['inspected_person'], record_data['responsible_team'],
                    record_data['assessment'], record_data['rectification_status'],
                    record_data['rectifier'], record_data['work_type'],
                    record_data['responsibility_location'], record_data['inspection_item'],
                    uid, filename
                ))
                total_imported += 1

            conn.commit()

            # 清理临时文件和session
            cleanup_temp_file()

            # 显示结果和记录日志
            total_rows = total_imported + total_skipped + total_skipped_no_permission
            msg_parts = []

            if total_imported > 0:
                msg_parts.append(f"成功导入 {total_imported} 条安全检查记录")
            if total_skipped > 0:
                msg_parts.append(f"{total_skipped} 条重复记录被跳过")
            if total_skipped_no_permission > 0:
                msg_parts.append(f"{total_skipped_no_permission} 条记录因无权限被跳过")

            if msg_parts:
                flash("、".join(msg_parts), "success" if total_imported > 0 else "warning")

            # 记录导入操作日志
            log_import_operation(
                module='safety',
                operation='import_confirm_duplicates',
                file_name=filename,
                total_rows=total_rows,
                success_rows=total_imported,
                failed_rows=0,
                skipped_rows=total_skipped + total_skipped_no_permission,
                import_details={
                    'imported': total_imported,
                    'skipped_duplicate': total_skipped,
                    'skipped_no_permission': total_skipped_no_permission,
                    'accessible_departments': len(accessible_dept_ids),
                    'duplicate_selections': len(emp_no_selections)
                }
            )

            return redirect(url_for("safety.records"))

        except Exception as e:
            error_msg = f"导入失败: {str(e)}"
            conn.rollback()
            flash(error_msg, "danger")

            # 记录失败的导入操作
            log_import_operation(
                module='safety',
                operation='import_confirm_duplicates',
                file_name=filename,
                total_rows=0,
                success_rows=0,
                failed_rows=0,
                skipped_rows=0,
                error_message=error_msg
            )

            return redirect(url_for('safety.confirm_duplicates'))

    # GET请求，展示选择页面
    # 获取部门信息
    conn = get_db()
    cur = conn.cursor()

    # 为每个重名人员获取部门信息
    duplicate_names_with_dept = {}

    for person_name, employees in pending_data['duplicate_names'].items():
        employees_with_dept = []

        for emp in employees:
            dept_name = "未分配部门"
            if emp['dept_id']:
                cur.execute("SELECT name FROM departments WHERE id = ?", (emp['dept_id'],))
                dept_row = cur.fetchone()
                if dept_row:
                    dept_name = dept_row[0]

            employees_with_dept.append({
                'emp_no': emp['emp_no'],
                'name': emp['name'],
                'dept_name': dept_name
            })

        duplicate_names_with_dept[person_name] = employees_with_dept

    return render_template(
        "safety_confirm_duplicates.html",
        title=f"确认重名人员 | {APP_TITLE}",
        filename=pending_data['filename'],
        duplicate_names=duplicate_names_with_dept,
        record_count=len(pending_data['records'])
    )


@safety_bp.route('/records')
@login_required
def records():
    """安全检查记录列表和导出"""
    # 使用统一的日期筛选器
    start_date, end_date = parse_date_filters('current_month')

    category_filter = request.args.get("category", "").strip()
    team_filter = request.args.get("team", "").strip()
    inspected_person_filter = request.args.get("inspected_person", "").strip()
    work_type_filter = request.args.get("work_type", "").strip()
    rectification_status_filter = request.args.get("rectification_status", "").strip()

    conn = get_db()
    cur = conn.cursor()

    # 获取当前用户角色
    user_id = session.get('user_id')
    cur.execute("SELECT role FROM users WHERE id = ?", (user_id,))
    row = cur.fetchone()
    user_role = row['role'] if row else 'user'

    # 使用部门过滤机制 - 通过inspected_person关联employees.name
    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return render_template(
            "safety_records.html",
            title=f"安全检查记录 | {APP_TITLE}",
            records=[],
            start_date=start_date or "",
            end_date=end_date or "",
            category_filter=category_filter,
            team_filter=team_filter,
            inspected_person_filter=inspected_person_filter,
            work_type_filter=work_type_filter,
            rectification_status_filter=rectification_status_filter,
            categories=[],
            teams=[],
            work_types=[],
            user_role=user_role,
        )

    # 构建查询 - 通过被检查人姓名关联employees表进行部门过滤
    placeholders = ','.join('?' * len(dept_ids))
    base_query = f"""
        SELECT sr.*
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
    """
    params = dept_ids.copy()

    # 应用日期筛选
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        base_query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)
    if category_filter:
        base_query += " AND sr.category LIKE ?"
        params.append(f"%{category_filter}%")
    if team_filter:
        base_query += " AND sr.responsible_team LIKE ?"
        params.append(f"%{team_filter}%")
    if inspected_person_filter:
        base_query += " AND sr.inspected_person LIKE ?"
        params.append(f"%{inspected_person_filter}%")
    if work_type_filter:
        base_query += " AND sr.work_type LIKE ?"
        params.append(f"%{work_type_filter}%")
    if rectification_status_filter:
        base_query += " AND sr.rectification_status LIKE ?"
        params.append(f"%{rectification_status_filter}%")

    base_query += " ORDER BY sr.inspection_date DESC, sr.id DESC"

    cur.execute(base_query, tuple(params))
    records = cur.fetchall()

    # 获取类别、车队和作业类型列表用于筛选
    cur.execute(f"""
        SELECT DISTINCT sr.category FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        ORDER BY sr.category
    """, tuple(dept_ids))
    categories = [row[0] for row in cur.fetchall() if row[0]]

    cur.execute(f"""
        SELECT DISTINCT sr.responsible_team FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        ORDER BY sr.responsible_team
    """, tuple(dept_ids))
    teams = [row[0] for row in cur.fetchall() if row[0]]

    cur.execute(f"""
        SELECT DISTINCT sr.work_type FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
          AND sr.work_type IS NOT NULL AND sr.work_type != ''
        ORDER BY sr.work_type
    """, tuple(dept_ids))
    work_types = [row[0] for row in cur.fetchall() if row[0]]

    return render_template(
        "safety_records.html",
        title=f"安全检查记录 | {APP_TITLE}",
        records=[dict(row) for row in records],
        start_date=start_date or "",
        end_date=end_date or "",
        category_filter=category_filter,
        team_filter=team_filter,
        inspected_person_filter=inspected_person_filter,
        work_type_filter=work_type_filter,
        rectification_status_filter=rectification_status_filter,
        categories=categories,
        teams=teams,
        work_types=work_types,
        user_role=user_role,
    )


@safety_bp.route('/analytics')
@login_required
def analytics():
    """安全统计分析和图表"""
    return render_template(
        "safety_analytics_new.html",
        title=f"安全统计 | {APP_TITLE}",
    )


@safety_bp.route('/export')
@login_required
def export():
    """导出安全检查记录到Excel"""
    # 使用统一的日期筛选器
    start_date, end_date = parse_date_filters('current_month')

    category_filter = request.args.get("category", "").strip()
    team_filter = request.args.get("team", "").strip()
    inspected_person_filter = request.args.get("inspected_person", "").strip()
    work_type_filter = request.args.get("work_type", "").strip()
    rectification_status_filter = request.args.get("rectification_status", "").strip()

    conn = get_db()
    cur = conn.cursor()

    # 使用部门过滤机制
    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        flash("无权限访问数据", "warning")
        return redirect(url_for("safety.records"))

    placeholders = ','.join('?' * len(dept_ids))
    base_query = f"""
        SELECT sr.*
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
    """
    params = dept_ids.copy()

    # 应用日期筛选
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        base_query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)
    if category_filter:
        base_query += " AND sr.category LIKE ?"
        params.append(f"%{category_filter}%")
    if team_filter:
        base_query += " AND sr.responsible_team LIKE ?"
        params.append(f"%{team_filter}%")
    if inspected_person_filter:
        base_query += " AND sr.inspected_person LIKE ?"
        params.append(f"%{inspected_person_filter}%")
    if work_type_filter:
        base_query += " AND sr.work_type LIKE ?"
        params.append(f"%{work_type_filter}%")
    if rectification_status_filter:
        base_query += " AND sr.rectification_status LIKE ?"
        params.append(f"%{rectification_status_filter}%")

    base_query += " ORDER BY sr.inspection_date DESC"

    cur.execute(base_query, tuple(params))
    rows = cur.fetchall()

    if not rows:
        flash("无数据可导出", "warning")
        return redirect(url_for("safety.records"))

    filename_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = os.path.join(EXPORT_DIR, f"安全检查记录_{filename_date}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "安全检查记录"

    headers = [
        "类别", "检查日期", "地点", "存在隐患和问题", "整改措施及其它意见",
        "整改期限", "被检查人", "责任车队", "考核情况", "整改情况",
        "整改人", "作业类型", "责任点位", "检查项目"
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["category"], row["inspection_date"], row["location"] or "",
            row["hazard_description"] or "", row["corrective_measures"] or "",
            row["deadline_date"] or "", row["inspected_person"] or "",
            row["responsible_team"] or "", row["assessment"] or "",
            row["rectification_status"] or "", row["rectifier"] or "",
            row["work_type"] or "", row["responsibility_location"] or "",
            row["inspection_item"] or ""
        ])

    wb.save(xlsx_path)
    return send_file(xlsx_path, as_attachment=True, download_name=os.path.basename(xlsx_path))


@safety_bp.route('/api/data')
@login_required
def api_data():
    """API端点，获取过滤后的安全检查数据（用于前端图表）"""
    conn = get_db()
    cur = conn.cursor()

    # 使用部门过滤机制
    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify([])

    placeholders = ','.join('?' * len(dept_ids))
    base_query = f"""
        SELECT sr.*
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
    """
    params = dept_ids.copy()

    start_date = request.args.get("start_date")
    if start_date:
        base_query += " AND sr.inspection_date >= ?"
        params.append(start_date)
    end_date = request.args.get("end_date")
    if end_date:
        base_query += " AND sr.inspection_date <= ?"
        params.append(end_date)

    category = request.args.get("category")
    if category:
        base_query += " AND sr.category LIKE ?"
        params.append(f"%{category}%")

    team = request.args.get("team")
    if team:
        base_query += " AND sr.responsible_team LIKE ?"
        params.append(f"%{team}%")

    base_query += " ORDER BY sr.inspection_date DESC"

    cur.execute(base_query, tuple(params))
    rows = cur.fetchall()

    data = [dict(row) for row in rows]
    return jsonify(data)


@safety_bp.route('/records/<int:record_id>/edit', methods=['POST'])
@role_required('manager')
def edit_record(record_id):
    """编辑安全检查记录（仅限部门管理员及以上权限）"""
    conn = get_db()
    cur = conn.cursor()

    # 获取表单数据
    category = request.form.get('category', '').strip()
    inspection_date = request.form.get('inspection_date', '').strip()
    location = request.form.get('location', '').strip()
    hazard_description = request.form.get('hazard_description', '').strip()
    corrective_measures = request.form.get('corrective_measures', '').strip()
    deadline_date = request.form.get('deadline_date', '').strip()
    inspected_person = request.form.get('inspected_person', '').strip()
    responsible_team = request.form.get('responsible_team', '').strip()
    assessment = request.form.get('assessment', '').strip()
    rectification_status = request.form.get('rectification_status', '').strip()
    rectifier = request.form.get('rectifier', '').strip()
    work_type = request.form.get('work_type', '').strip()
    responsibility_location = request.form.get('responsibility_location', '').strip()
    inspection_item = request.form.get('inspection_item', '').strip()

    # 验证必填字段
    if not category or not inspection_date:
        flash('类别和检查日期为必填项', 'warning')
        return redirect(url_for('safety.records'))

    try:
        # 更新记录
        cur.execute("""
            UPDATE safety_inspection_records
            SET category = ?, inspection_date = ?, location = ?,
                hazard_description = ?, corrective_measures = ?,
                deadline_date = ?, inspected_person = ?,
                responsible_team = ?, assessment = ?,
                rectification_status = ?, rectifier = ?,
                work_type = ?, responsibility_location = ?,
                inspection_item = ?
            WHERE id = ?
        """, (category, inspection_date, location, hazard_description,
              corrective_measures, deadline_date, inspected_person,
              responsible_team, assessment, rectification_status,
              rectifier, work_type, responsibility_location,
              inspection_item, record_id))

        conn.commit()
        flash('安全检查记录已更新', 'success')
    except Exception as e:
        flash(f'更新失败: {e}', 'danger')

    return redirect(url_for('safety.records'))


@safety_bp.route('/records/<int:record_id>/delete', methods=['POST'])
@role_required('manager')
def delete_record(record_id):
    """删除安全检查记录（仅限部门管理员及以上权限）"""
    conn = get_db()
    cur = conn.cursor()

    try:
        # 删除记录
        cur.execute("DELETE FROM safety_inspection_records WHERE id = ?", (record_id,))
        conn.commit()
        flash('安全检查记录已删除', 'success')
    except Exception as e:
        flash(f'删除失败: {e}', 'danger')

    return redirect(url_for('safety.records'))


@safety_bp.route('/records/batch-delete', methods=['POST'])
@role_required('manager')
def batch_delete_records():
    """批量删除安全检查记录（仅限部门管理员及以上权限）"""
    conn = get_db()
    cur = conn.cursor()

    record_ids = request.form.getlist('record_ids')

    if not record_ids:
        flash('未选择要删除的记录', 'warning')
        return redirect(url_for('safety.records'))

    try:
        # 批量删除记录
        placeholders = ','.join('?' * len(record_ids))
        cur.execute(f"DELETE FROM safety_inspection_records WHERE id IN ({placeholders})", record_ids)
        conn.commit()
        flash(f'成功删除 {len(record_ids)} 条安全检查记录', 'success')
    except Exception as e:
        flash(f'批量删除失败: {e}', 'danger')

    return redirect(url_for('safety.records'))


# ==================== 安全数据分析 API ====================

import re

def extract_score_from_assessment(assessment):
    """从考核情况中提取分值"""
    if not assessment:
        return 0

    # 过滤正面评价
    positive_keywords = ['继续发扬', '正常', '良好', '优秀', '表扬']
    for keyword in positive_keywords:
        if keyword in assessment:
            return 0

    # 过滤直接扣钱的情况（扣100元等）- 这些不应该被统计为违规扣分
    money_keywords = ['元', '钱', '¥', '￥', 'RMB', 'rmb']
    for keyword in money_keywords:
        if keyword in assessment:
            return 0

    # 提取数字（支持小数）
    numbers = re.findall(r'\d+\.?\d*', assessment)
    if numbers:
        return float(numbers[-1])  # 取最后一个数字

    # 如果没有数字但不是正面评价，默认赋值1
    return 1


@safety_bp.route('/api/analytics/severity-distribution')
@login_required
def api_analytics_severity_distribution():
    """图表A：问题严重度分布（饼图数据）"""
    conn = get_db()
    cur = conn.cursor()

    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify([])

    # 使用统一的日期筛选函数
    start_date, end_date = parse_date_filters('current_month')

    placeholders = ','.join('?' * len(dept_ids))
    query = f"""
        SELECT sr.assessment
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        AND sr.assessment IS NOT NULL
        AND sr.assessment != ''
    """
    params = list(dept_ids)

    # 使用统一的日期条件构建函数
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # 数据预处理：提取分值并统计
    score_counts = {}
    for row in rows:
        score = extract_score_from_assessment(row[0])
        if score > 0:  # 只统计有效扣分
            # 将浮点数转换为整数(如果是整数值),避免1.0和1分开统计
            if score == int(score):
                score_key = f"{int(score)}分"
            else:
                score_key = f"{score}分"
            score_counts[score_key] = score_counts.get(score_key, 0) + 1

    # 转换为ECharts饼图格式
    result = [{"name": name, "value": count} for name, count in score_counts.items()]

    return jsonify(result)


@safety_bp.route('/api/analytics/daily-trend')
@login_required
def api_analytics_daily_trend():
    """图表B：每日违规趋势（双轴图数据）"""
    conn = get_db()
    cur = conn.cursor()

    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify({"dates": [], "counts": [], "scores": []})

    # 使用统一的日期筛选函数
    start_date, end_date = parse_date_filters('current_month')

    placeholders = ','.join('?' * len(dept_ids))
    query = f"""
        SELECT sr.inspection_date, sr.assessment
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        AND sr.assessment IS NOT NULL
        AND sr.assessment != ''
    """
    params = list(dept_ids)

    # 使用统一的日期条件构建函数
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)

    query += " ORDER BY sr.inspection_date ASC"

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # 按日期聚合
    daily_data = {}
    for row in rows:
        date = row[0]
        score = extract_score_from_assessment(row[1])
        if score > 0:
            if date not in daily_data:
                daily_data[date] = {"count": 0, "total_score": 0}
            daily_data[date]["count"] += 1
            daily_data[date]["total_score"] += score

    # 排序并转换格式
    dates = sorted(daily_data.keys())
    counts = [daily_data[date]["count"] for date in dates]
    # 规范化分值显示：整数显示为整数,小数保留一位
    scores = [
        int(daily_data[date]["total_score"]) if daily_data[date]["total_score"] == int(daily_data[date]["total_score"])
        else round(daily_data[date]["total_score"], 1)
        for date in dates
    ]

    return jsonify({
        "dates": dates,
        "counts": counts,
        "scores": scores
    })


@safety_bp.route('/api/analytics/top-loss-items')
@login_required
def api_analytics_top_loss_items():
    """图表C：高频失分项目 Top 10（横向条形图数据）"""
    conn = get_db()
    cur = conn.cursor()

    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify([])

    # 使用统一的日期筛选函数
    start_date, end_date = parse_date_filters('current_month')

    placeholders = ','.join('?' * len(dept_ids))
    query = f"""
        SELECT sr.inspection_item, sr.assessment
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        AND sr.inspection_item IS NOT NULL
        AND sr.inspection_item != ''
        AND sr.assessment IS NOT NULL
        AND sr.assessment != ''
    """
    params = list(dept_ids)

    # 使用统一的日期条件构建函数
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # 按检查项目聚合
    item_scores = {}
    for row in rows:
        item = row[0]
        score = extract_score_from_assessment(row[1])
        if score > 0:
            item_scores[item] = item_scores.get(item, 0) + score

    # 排序并取Top 10
    sorted_items = sorted(item_scores.items(), key=lambda x: x[1], reverse=True)[:10]

    # 转换为ECharts横向条形图格式（Y轴是项目名，X轴是分值）
    # 规范化分值显示：整数显示为整数,小数保留一位
    result = {
        "items": [item[0] for item in sorted_items],
        "scores": [int(item[1]) if item[1] == int(item[1]) else round(item[1], 1) for item in sorted_items]
    }

    return jsonify(result)


@safety_bp.route('/api/analytics/personnel-risk-matrix')
@login_required
def api_analytics_personnel_risk_matrix():
    """图表D：人员风险矩阵（散点图数据）"""
    conn = get_db()
    cur = conn.cursor()

    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify([])

    # 使用统一的日期筛选函数
    start_date, end_date = parse_date_filters('current_month')

    placeholders = ','.join('?' * len(dept_ids))
    query = f"""
        SELECT
            sr.inspected_person,
            sr.responsible_team,
            sr.assessment
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        AND sr.inspected_person IS NOT NULL
        AND sr.inspected_person != ''
        AND sr.assessment IS NOT NULL
        AND sr.assessment != ''
    """
    params = list(dept_ids)

    # 使用统一的日期条件构建函数
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # 按人员聚合
    personnel_data = {}
    for row in rows:
        person = row[0]
        team = row[1] or "未知"
        score = extract_score_from_assessment(row[2])

        if score > 0:
            if person not in personnel_data:
                personnel_data[person] = {
                    "team": team,
                    "count": 0,
                    "total_score": 0
                }
            personnel_data[person]["count"] += 1
            personnel_data[person]["total_score"] += score

    # 转换为散点图格式
    result = []
    for person, data in personnel_data.items():
        # 规范化分值显示：整数显示为整数,小数保留一位
        total_score = data["total_score"]
        normalized_score = int(total_score) if total_score == int(total_score) else round(total_score, 1)

        result.append({
            "name": person,
            "team": data["team"],
            "value": [data["count"], normalized_score]  # [X轴违规次数, Y轴累计扣分]
        })

    return jsonify(result)


@safety_bp.route('/api/analytics/top-contributors')
@login_required
def api_analytics_top_contributors():
    """图表E：问题发现能手榜 Top 10（柱状图数据）"""
    conn = get_db()
    cur = conn.cursor()

    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify({"names": [], "counts": []})

    # 使用统一的日期筛选函数
    start_date, end_date = parse_date_filters('current_month')

    placeholders = ','.join('?' * len(dept_ids))
    query = f"""
        SELECT sr.rectifier
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        AND sr.rectifier IS NOT NULL
        AND sr.rectifier != ''
    """
    params = list(dept_ids)

    # 使用统一的日期条件构建函数
    date_conditions, date_params = build_date_filter_sql('sr.inspection_date', start_date, end_date)
    if date_conditions:
        query += " AND " + " AND ".join(date_conditions)
        params.extend(date_params)

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # 统计每个整改人的问题数量
    contributor_counts = {}
    for row in rows:
        rectifier = row[0]
        contributor_counts[rectifier] = contributor_counts.get(rectifier, 0) + 1

    # 排序并取Top 10
    sorted_contributors = sorted(contributor_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    result = {
        "names": [item[0] for item in sorted_contributors],
        "counts": [item[1] for item in sorted_contributors]
    }

    return jsonify(result)


@safety_bp.route('/api/analytics/severity-drilldown')
@login_required
def api_analytics_severity_drilldown():
    """问题严重度下钻API - 获取特定分数段的问题/隐患列表"""
    conn = get_db()
    cur = conn.cursor()

    dept_ids = get_accessible_department_ids()
    if not dept_ids:
        return jsonify({"canDrilldown": False, "message": "无权限访问数据"})

    # 获取分数参数
    score_str = request.args.get('score', '').strip()
    if not score_str:
        return jsonify({"canDrilldown": False, "message": "缺少分数参数"})

    # 解析分数（去除"分"字）
    try:
        score = float(score_str.replace('分', ''))
    except ValueError:
        return jsonify({"canDrilldown": False, "message": "分数格式错误"})

    placeholders = ','.join('?' * len(dept_ids))
    query = f"""
        SELECT
            sr.id,
            sr.inspection_date,
            sr.inspected_person,
            sr.responsible_team,
            sr.hazard_description,
            sr.corrective_measures,
            sr.assessment,
            sr.rectification_status,
            sr.location,
            sr.inspection_item
        FROM safety_inspection_records sr
        LEFT JOIN employees e ON sr.inspected_person = e.name
        WHERE (e.department_id IN ({placeholders}) OR sr.inspected_person IS NULL OR sr.inspected_person = '')
        AND sr.assessment IS NOT NULL
        AND sr.assessment != ''
    """
    params = list(dept_ids)

    # 添加日期筛选
    start_date = request.args.get('start_date')
    if start_date:
        query += " AND sr.inspection_date >= ?"
        params.append(start_date)
    end_date = request.args.get('end_date')
    if end_date:
        query += " AND sr.inspection_date <= ?"
        params.append(end_date)

    query += " ORDER BY sr.inspection_date DESC"

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # 筛选出对应分数的问题记录
    problem_records = []
    for row in rows:
        record_score = extract_score_from_assessment(row[6])  # assessment 在索引6

        # 判断是否属于该分数段
        if record_score == score:
            problem_records.append({
                "id": row[0],
                "date": row[1],
                "inspectedPerson": row[2] or "未知",
                "team": row[3] or "未知",
                "hazardDescription": row[4] or "",
                "correctiveMeasures": row[5] or "",
                "assessment": row[6] or "",
                "rectificationStatus": row[7] or "待整改",
                "location": row[8] or "",
                "inspectionItem": row[9] or "",
                "score": int(record_score) if record_score == int(record_score) else round(record_score, 1)
            })

    # 下钻逻辑判断
    problem_count = len(problem_records)

    # 1分以下且问题过多（超过30条）不支持下钻
    if score <= 1 and problem_count > 30:
        return jsonify({
            "canDrilldown": False,
            "message": f"该分数段问题过多（{problem_count}条），不支持下钻查看",
            "problemCount": problem_count
        })

    # 2分以上必须支持下钻
    # 或者1分以下但问题不多的情况也支持下钻
    return jsonify({
        "canDrilldown": True,
        "score": score,
        "scoreLabel": score_str,
        "problemCount": problem_count,
        "problems": problem_records,
        "message": f"找到 {problem_count} 条问题记录"
    })
